import base64
import os
import re
import socket
import uuid
import json as json_mod
import traceback
import queue
from datetime import datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, send_file, Response
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
from werkzeug.utils import secure_filename
import config as app_config
from config import (
    UPLOAD_FOLDER, DEEPSEEK_API_KEY, DEEPSEEK_BASE_URL,
    SECRET_KEY, CORS_ORIGINS, BAIDU_APP_KEY, BAIDU_SECRET_KEY,
    ERNIE_API_KEY, ERNIE_BASE_URL, RAG_DB_PATH, PDF_OUTPUT_DIR
)
from auth import create_token, require_session, revoke_session
from core.model_registry import init_providers, get_provider, get_default_provider, list_available_providers
from services.ocr_result_utils import normalize_reusable_ocr_result
from services.resume_preview_service import get_resume_preview_summary
from services.resume_text_extraction import (
    ResumeExtractionError,
    ResumeOcrUnavailableError,
    ensure_supported_resume_extension,
    extract_resume_content,
    unwrap_resume_text,
)
from services.career_planning_service import CareerPlanningService
from services.career_planning_docs import CareerPlanningDocumentRepository
from services.interview_turn_service import InterviewTurnService
from utils.safe_log import configure_stdio, safe_log
from monitoring.routes import monitoring_bp, set_data_client_provider

configure_stdio()

# Keep existing module-level diagnostics safe on Windows GBK/CP936 stdout.
print = safe_log

# 直接导入 OCR 工具（不再依赖 Agent 自主决策）
try:
    from core.tools.ocr_processing import perform_ocr, perform_ocr_full
    OCR_AVAILABLE = True
except Exception as e:
    print(f"[WARN] Import OCR tools failed: {e}")
    OCR_AVAILABLE = False

# 创建 Flask app 实例
app = Flask(__name__)
app.secret_key = SECRET_KEY
CORS(app, origins=CORS_ORIGINS, supports_credentials=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
set_data_client_provider(lambda: data_client if STORAGE_AVAILABLE else None)
app.register_blueprint(monitoring_bp)

def _reload_runtime_config_state() -> dict:
    global DEEPSEEK_API_KEY, DEEPSEEK_BASE_URL
    global ERNIE_API_KEY, ERNIE_BASE_URL
    global BAIDU_APP_KEY, BAIDU_SECRET_KEY

    snapshot = app_config.reload_runtime_settings()

    DEEPSEEK_API_KEY = app_config.DEEPSEEK_API_KEY
    DEEPSEEK_BASE_URL = app_config.DEEPSEEK_BASE_URL
    ERNIE_API_KEY = app_config.ERNIE_API_KEY
    ERNIE_BASE_URL = app_config.ERNIE_BASE_URL
    BAIDU_APP_KEY = app_config.BAIDU_APP_KEY
    BAIDU_SECRET_KEY = app_config.BAIDU_SECRET_KEY

    init_providers(
        deepseek_api_key=DEEPSEEK_API_KEY,
        deepseek_base_url=DEEPSEEK_BASE_URL,
        ernie_api_key=ERNIE_API_KEY,
        ernie_base_url=ERNIE_BASE_URL,
    )
    return snapshot


_reload_runtime_config_state()

# 加载岗位数据（启动时一次性读取 Excel）
_positions: list[str] = []
def _decode_legacy_mojibake(value: str) -> str:
    return base64.b64decode(value).decode("utf-8")


_LEGACY_GARBLED_SENIOR_FRONTEND_TITLE_A = _decode_legacy_mojibake(
    "5qWC5qiG6aqC5Zus5bSc5ayl5LyC54C14oaC4oCT6Za45qyS5Z615rW85oSu57Kz54Cj7oeO"
)
_LEGACY_GARBLED_SENIOR_FRONTEND_TITLE_B = _decode_legacy_mojibake(
    "5qWg5qiH5rW36Y2Z5qiA5Zus7oGs5a+u7pKC5bGd6Y+C7oCX5aCu55KHz7flmp8="
)
_garbled_job_title_map = {
    _LEGACY_GARBLED_SENIOR_FRONTEND_TITLE_A: "高级前端开发工程师",
    _LEGACY_GARBLED_SENIOR_FRONTEND_TITLE_B: "高级前端开发工程师",
}
try:
    import openpyxl
    if os.path.exists(RAG_DB_PATH):
        _wb = openpyxl.load_workbook(RAG_DB_PATH, read_only=True)
        _ws = _wb[_wb.sheetnames[0]]
        _seen = set()
        for row in _ws.iter_rows(min_row=2, values_only=True):
            name = row[1]  # B 列：岗位名称
            if name and str(name).strip() and str(name).strip() not in _seen:
                _seen.add(str(name).strip())
                _positions.append(str(name).strip())
        _wb.close()
        print(f"[OK] Loaded {len(_positions)} positions")
    else:
        print(f"[WARN] RAG database file not found: {RAG_DB_PATH}")
except Exception as e:
    print(f"[WARN] Failed to load positions: {e}")

# 导入核心模块
def normalize_job_title(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return _garbled_job_title_map.get(text, text)


def normalize_session_info(session_info):
    if not isinstance(session_info, dict):
        return session_info
    normalized = dict(session_info)
    if "position" in normalized:
        normalized["position"] = normalize_job_title(normalized.get("position"))
    return normalized


def normalize_session_list(sessions):
    return [normalize_session_info(session) for session in (sessions or [])]


def _get_server_host() -> str:
    host = str(os.getenv("PROVIEW_API_HOST", "127.0.0.1")).strip()
    return host or "127.0.0.1"


def _get_server_port() -> int:
    raw_port = str(os.getenv("PROVIEW_API_PORT") or os.getenv("PORT") or "5000").strip()
    try:
        port = int(raw_port)
    except ValueError:
        print(f"[WARN] Invalid PROVIEW_API_PORT={raw_port!r}, fallback to 5000")
        return 5000
    if 1 <= port <= 65535:
        return port
    print(f"[WARN] Out-of-range PROVIEW_API_PORT={raw_port!r}, fallback to 5000")
    return 5000


def _is_truthy_env(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _should_enable_debug_mode() -> bool:
    for key in ("PROVIEW_DEBUG", "FLASK_DEBUG"):
        raw_value = os.getenv(key)
        if raw_value is not None and str(raw_value).strip():
            return _is_truthy_env(raw_value)

    if _is_truthy_env(os.getenv("PROVIEW_DESKTOP_MODE")):
        return False

    return True


def _is_api_request() -> bool:
    return str(request.path or "").startswith("/api/")


@app.errorhandler(HTTPException)
def _handle_api_http_exception(exc):
    if not _is_api_request():
        return exc

    return jsonify({
        "status": "error",
        "message": exc.description or exc.name,
    }), exc.code


@app.errorhandler(Exception)
def _handle_api_unexpected_exception(exc):
    traceback.print_exc()

    if not _is_api_request():
        return "Internal Server Error", 500

    return jsonify({
        "status": "error",
        "message": "服务器内部错误，请稍后重试。",
    }), 500


def _build_server_url(host: str, port: int) -> str:
    display_host = "127.0.0.1" if host in {"0.0.0.0", "::"} else host
    return f"http://{display_host}:{port}"


def _assert_server_bindable(host: str, port: int) -> None:
    probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        probe.bind((host, port))
    except OSError as exc:
        error_code = getattr(exc, "winerror", None) or getattr(exc, "errno", None)
        if error_code in {10013, 13}:
            reason = "访问被拒绝，常见原因是端口被其他程序独占，或被 Windows 保留。"
        elif error_code in {10048, 98}:
            reason = "端口已被其他进程占用。"
        else:
            reason = str(exc)
        raise RuntimeError(
            f"无法绑定 {host}:{port}。{reason} "
            "请关闭占用该端口的程序，或在 backend/.env 中设置 PROVIEW_API_PORT=其它端口。"
        ) from exc
    finally:
        probe.close()


def _save_uploaded_resume(file_storage):
    raw_name = str(getattr(file_storage, "filename", "") or "").strip()
    ext = ensure_supported_resume_extension(raw_name)
    filename = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file_storage.save(file_path)
    return filename, file_path


def _extract_resume_payload(file_path: str, *, include_images: bool = False) -> dict:
    return extract_resume_content(
        file_path,
        include_images=include_images,
        ocr_available=OCR_AVAILABLE,
        ocr_text_loader=perform_ocr if OCR_AVAILABLE else None,
        ocr_full_loader=perform_ocr_full if OCR_AVAILABLE else None,
        use_preprocessing=True,
    )


MIN_RESUME_TEXT_LENGTH = 50


def _cleanup_local_resume_file(file_path: str) -> None:
    path = str(file_path or "").strip()
    if not path:
        return
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass


def _has_usable_resume_text(value: object, minimum_chars: int = MIN_RESUME_TEXT_LENGTH) -> bool:
    return len(str(value or "").strip()) >= minimum_chars


def _build_reused_resume_payload(raw_text: object, *, source_label: str) -> dict:
    raw_value = str(raw_text or "").strip()
    reusable_text = normalize_reusable_ocr_result(raw_text)

    if not reusable_text and _has_usable_resume_text(raw_value):
        clean_raw_text = unwrap_resume_text(raw_value)
        if _has_usable_resume_text(clean_raw_text):
            reusable_text = f"【解析成功】已复用{source_label}内容\n\n以下是提取的内容:\n\n{clean_raw_text}"

    clean_text = unwrap_resume_text(reusable_text)

    if not reusable_text:
        return {
            "success": False,
            "mode": "reused_text",
            "source_label": source_label,
            "text": "",
            "reusable_text": "",
            "raw_text": raw_value,
            "images": {},
            "error_message": f"{source_label}内容不可复用，请重新上传简历。",
        }

    if not _has_usable_resume_text(clean_text):
        return {
            "success": False,
            "mode": "reused_text",
            "source_label": source_label,
            "text": "",
            "reusable_text": reusable_text,
            "raw_text": reusable_text,
            "images": {},
            "error_message": f"{source_label}内容过少或为空，请重新上传简历。",
        }

    return {
        "success": True,
        "mode": "reused_text",
        "source_label": source_label,
        "text": clean_text,
        "reusable_text": reusable_text,
        "raw_text": reusable_text,
        "images": {},
        "error_message": "",
    }


def _get_current_user_id_from_auth_header():
    user_info = _get_current_user_info()
    if user_info and "id" in user_info:
        return user_info["id"]
    return None


def _get_local_user_name() -> str:
    value = str(getattr(app_config, "LOCAL_USER_NAME", "") or "").strip()
    return value or "本地用户"


def _apply_local_user_alias(user_info):
    if not isinstance(user_info, dict):
        return None

    alias = _get_local_user_name()
    normalized = dict(user_info)
    normalized["username"] = alias
    normalized["display_name"] = alias
    return normalized


def _get_current_user_info():
    if not STORAGE_AVAILABLE or not data_client:
        return None

    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        user_info = data_client.get_user(auth_header[7:])
        if user_info and "id" in user_info:
            return _apply_local_user_alias(user_info)

    local_user = data_client.get_or_create_local_user(_get_local_user_name())
    if local_user and "id" in local_user:
        return _apply_local_user_alias(local_user)
    return None


def _parse_career_plan_generate_payload(payload):
    if not isinstance(payload, dict):
        raise ValueError("请求体必须是 JSON 对象")

    target_role = str(payload.get("target_role") or "").strip()
    career_goal = str(payload.get("career_goal") or "").strip()

    try:
        horizon_months = int(payload.get("horizon_months") or 6)
    except (TypeError, ValueError) as exc:
        raise ValueError("horizon_months 必须是整数") from exc

    refresh = bool(payload.get("refresh", False))
    return target_role, career_goal, horizon_months, refresh


def _build_history_quota(user_id):
    saved_count = data_client.count_user_sessions(user_id) if STORAGE_AVAILABLE and data_client else 0
    return {
        "saved_count": saved_count,
        "max_saved": None,
        "remaining": None,
        "can_save": True,
    }


def _parse_iso_datetime(value):
    raw = str(value or "").strip()
    if not raw:
        return None

    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _is_new_user(user_info):
    if not user_info or "id" not in user_info:
        return False

    created_at = _parse_iso_datetime(user_info.get("created_at"))
    if not created_at:
        return False

    return datetime.now(timezone.utc) - created_at <= NEW_USER_HISTORY_GRACE_PERIOD


def _serialize_resume_library_record(record):
    payload = dict(record or {})
    preview_summary = get_resume_preview_summary(payload.get("file_path") or "", payload.get("file_name") or "")
    payload["file_kind"] = preview_summary["file_kind"]
    payload["preview_page_count"] = preview_summary["preview_page_count"]
    payload["can_preview"] = preview_summary["has_preview"]
    payload["preview_cover_url"] = (
        f"/api/my-resumes/{payload['id']}/preview/1" if preview_summary["has_preview"] else ""
    )
    payload["preview_image_urls"] = [
        f"/api/my-resumes/{payload['id']}/preview/{index + 1}"
        for index in range(preview_summary["preview_page_count"])
    ]
    return payload


AGENT_AVAILABLE = False
STORAGE_AVAILABLE = False
STORAGE_STATUS = {
    "connected": False,
    "mode": None,
    "url": None,
    "db_error": None,
    "fallback_reason": None,
}
STORAGE_INIT_ERROR = None
data_client = None
career_planning_service = None
career_planning_docs = CareerPlanningDocumentRepository()
turn_service = InterviewTurnService()
NEW_USER_HISTORY_GRACE_PERIOD = timedelta(minutes=30)


def _refresh_storage_status():
    global STORAGE_AVAILABLE, STORAGE_STATUS

    if not data_client:
        STORAGE_AVAILABLE = False
        STORAGE_STATUS = {
            "connected": False,
            "mode": None,
            "url": None,
            "db_error": STORAGE_INIT_ERROR or "storage client not initialized",
            "fallback_reason": None,
        }
        return STORAGE_STATUS

    health = data_client.health() or {}
    capabilities = data_client.storage_capabilities() if hasattr(data_client, "storage_capabilities") else {}
    STORAGE_AVAILABLE = bool(health.get("db_ok"))
    STORAGE_STATUS = {
        "connected": STORAGE_AVAILABLE,
        "mode": health.get("mode", getattr(data_client, "mode", None)),
        "url": health.get("db_url") or health.get("url"),
        "db_error": health.get("db_error"),
        "fallback_reason": health.get("fallback_reason"),
        "capabilities": capabilities,
    }
    return STORAGE_STATUS

try:
    from core.langchain_agent import LangChainInterviewAgent
    AGENT_AVAILABLE = True
except Exception as e:
    print(f"[WARN] Import langchain_agent failed: {e}")

# Supabase/Postgres 存储客户端
try:
    from data_client import DataServiceClient
    data_client = DataServiceClient()
    turn_service.set_data_client(data_client)
    _storage = _refresh_storage_status()
    if STORAGE_AVAILABLE:
        print(f"[storage] connected via {_storage.get('mode')}: {_storage.get('url')}")
    else:
        print(f"[storage] unavailable: {_storage.get('db_error')}")
    try:
        career_planning_service = CareerPlanningService(data_client)
        print(f"[career] schema ready: {career_planning_service.health()}")
    except Exception as career_error:
        career_planning_service = None
        print(f"[career] unavailable: {career_error}")
except Exception as e:
    STORAGE_INIT_ERROR = str(e)
    print(f"[storage] unavailable: {e}")
    data_client = None
    career_planning_service = None
    _refresh_storage_status()

# 导入简历分析模块
ANALYZER_AVAILABLE = False
try:
    from core.resume_analyzer import ResumeAnalyzer
    ANALYZER_AVAILABLE = True
except Exception as e:
    print(f"[WARN] Import ResumeAnalyzer failed: {e}")

# 导入百度语音模块
SPEECH_AVAILABLE = False
_speech_client = None


def _refresh_speech_client() -> None:
    global SPEECH_AVAILABLE, _speech_client
    SPEECH_AVAILABLE = False
    _speech_client = None

    try:
        if BAIDU_APP_KEY and BAIDU_SECRET_KEY:
            from core.baidu_speech import BaiduSpeechClient
            _speech_client = BaiduSpeechClient(
                app_key=BAIDU_APP_KEY, secret_key=BAIDU_SECRET_KEY
            )
            SPEECH_AVAILABLE = True
            print("[OK] Baidu speech module loaded")
        else:
            print("[WARN] BAIDU_APP_KEY / BAIDU_SECRET_KEY not configured, speech unavailable")
    except Exception as e:
        print(f"[WARN] Import Baidu speech module failed: {e}")


_refresh_speech_client()

# Per-session agent 管理
_agents: dict[str, object] = {}
_chat_stop_flags: dict[str, bool] = {}
_session_models: dict[str, str] = {}  # session_id → model provider key
_session_trace_contexts: dict[str, dict] = {}
_session_context_checkpoints: dict[str, dict] = {}
CONTEXT_COMPACTION_BUDGET_TOKENS = 8000
CONTEXT_COMPACTION_TRIGGER_RATIO = 0.6
CONTEXT_COMPACTION_MIN_TURNS = 6

# Per-session 评估观察者
try:
    from core.eval_observer import EvalObserver
    _observers: dict[str, "EvalObserver"] = {}
    OBSERVER_AVAILABLE = True
except Exception as _e:
    print(f"[WARN] Import EvalObserver failed: {_e}")
    _observers = {}
    OBSERVER_AVAILABLE = False


def get_agent(session_id: str, model_provider: str = ""):
    """获取或创建 session 对应的 agent 实例"""
    if session_id in _agents:
        return _agents[session_id]
    if not AGENT_AVAILABLE:
        return None

    # 根据前端选择的模型提供商获取配置
    provider = get_provider(model_provider) if model_provider else None
    if not provider or not provider.available:
        provider = get_default_provider()

    agent = LangChainInterviewAgent(
        api_key=provider.api_key,
        base_url=provider.base_url,
        model=provider.model,
        temperature=0.7,
        verbose=False
    )
    _agents[session_id] = agent
    _session_models[session_id] = provider.key

    # 创建评估观察者
    if OBSERVER_AVAILABLE and hasattr(agent, 'llm_client') and agent.llm_client:
        observer = EvalObserver(
            session_id=session_id,
            llm_client=agent.llm_client,
            data_client=data_client if STORAGE_AVAILABLE else None,
        )
        # 设置 SSE 推送回调（将在 chat-stream 中绑定到具体 SSE 响应）
        _observers[session_id] = observer

    return agent


def _register_session_trace_context(
    *,
    session_id: str,
    user_id=None,
    job_title: str = "",
    interview_type: str = "",
    difficulty: str = "",
    style: str = "",
    has_resume: bool = False,
    feature_vad: bool = False,
    feature_deep: bool = False,
):
    provider_key = _session_models.get(session_id, "unknown")
    provider = get_provider(provider_key)
    _session_trace_contexts[session_id] = {
        "user_id": str(user_id) if user_id is not None else "",
        "model_provider": provider_key,
        "model_name": provider.model if provider else "",
        "model_label": provider.label if provider else provider_key,
        "job_title": normalize_job_title(job_title),
        "interview_type": str(interview_type or ""),
        "difficulty": str(difficulty or ""),
        "interview_style": str(style or ""),
        "has_resume": bool(has_resume),
        "feature_vad": bool(feature_vad),
        "feature_deep": bool(feature_deep),
        "agent_role": "interviewer",
        "context_version": 1,
    }


def _build_langfuse_trace_context(
    session_id: str,
    *,
    interaction_type: str,
    trace_name: str = "",
    extra_metadata: dict | None = None,
) -> dict:
    base = dict(_session_trace_contexts.get(session_id) or {})
    provider_key = base.get("model_provider") or _session_models.get(session_id, "unknown")
    provider = get_provider(provider_key)
    if provider:
        base.setdefault("model_name", provider.model)
        base.setdefault("model_label", provider.label)
    base.setdefault("model_provider", provider_key)
    base.setdefault("agent_role", "interviewer")
    base.setdefault("context_version", 1)

    metadata = {
        "proview_schema_version": "trace_context_v1",
        "interaction_type": interaction_type,
        "agent_role": base.get("agent_role"),
        "model_provider": base.get("model_provider"),
        "model_name": base.get("model_name"),
        "model_label": base.get("model_label"),
        "job_title": base.get("job_title"),
        "interview_type": base.get("interview_type"),
        "difficulty": base.get("difficulty"),
        "interview_style": base.get("interview_style"),
        "has_resume": base.get("has_resume"),
        "feature_vad": base.get("feature_vad"),
        "feature_deep": base.get("feature_deep"),
        "context_version": base.get("context_version"),
    }
    if isinstance(extra_metadata, dict):
        for key, value in extra_metadata.items():
            if key not in {"resume_text", "ocr_raw_text", "job_requirements", "token"}:
                metadata[str(key)] = value

    tags = [
        "proview",
        "interview",
        f"interaction:{_trace_tag_value(interaction_type)}",
        f"model:{_trace_tag_value(provider_key)}",
        f"role:{_trace_tag_value(base.get('agent_role'))}",
    ]

    return {
        "session_id": session_id,
        "user_id": base.get("user_id"),
        "trace_name": trace_name or f"proview.interview.{interaction_type}",
        "tags": tags,
        "metadata": {key: value for key, value in metadata.items() if value not in (None, "")},
    }


def _trace_tag_value(value) -> str:
    text = str(value or "unknown").strip().lower()
    text = re.sub(r"[^a-z0-9_.:-]+", "-", text)
    return text.strip("-") or "unknown"


def _clear_session_runtime(session_id: str):
    _agents.pop(session_id, None)
    _session_models.pop(session_id, None)
    _session_trace_contexts.pop(session_id, None)
    _session_context_checkpoints.pop(session_id, None)
    _chat_stop_flags.pop(session_id, None)
    if turn_service:
        turn_service.clear_session(session_id)


def _append_visible_message(session_id: str, role: str, content: str):
    if not (STORAGE_AVAILABLE and data_client and content):
        return None
    _ensure_turn_service_client()
    message = turn_service.append_message(session_id, role, content) if turn_service else None
    if message:
        return message
    try:
        if data_client.save_message(session_id, role, content):
            return {"id": None, "session_id": session_id, "role": role, "content": content, "timestamp": ""}
    except Exception as exc:
        print(f"[app] save visible message failed: {exc}")
    return None


def _get_session_interview_metadata(session_id: str) -> dict:
    metadata = {}
    trace_context = _session_trace_contexts.get(session_id) or {}
    if trace_context:
        metadata.update({
            "difficulty": trace_context.get("difficulty") or "",
            "interview_type": trace_context.get("interview_type") or "",
            "style": trace_context.get("interview_style") or "",
            "job_title": trace_context.get("job_title") or "",
        })

    if STORAGE_AVAILABLE and data_client:
        try:
            session_info = data_client.get_session_info(session_id) or {}
            session_metadata = session_info.get("metadata") or {}
            metadata.setdefault("job_title", session_info.get("position") or "")
            if not metadata.get("difficulty"):
                metadata["difficulty"] = session_metadata.get("diff") or session_metadata.get("difficulty") or ""
            if not metadata.get("interview_type"):
                metadata["interview_type"] = session_metadata.get("type") or session_metadata.get("interview_type") or ""
            if not metadata.get("style"):
                metadata["style"] = session_info.get("interview_style") or session_metadata.get("style") or ""
        except Exception:
            pass

    return metadata


def _create_pending_turn(
    session_id: str,
    *,
    question_message,
    question_text: str,
    source: str = "interviewer_llm",
    question_type: str = "followup",
    difficulty: str = "",
    rag_candidates: list | None = None,
):
    if not (STORAGE_AVAILABLE and data_client and turn_service):
        return None
    _ensure_turn_service_client()
    session_metadata = _get_session_interview_metadata(session_id)
    return turn_service.create_pending_turn(
        session_id,
        question_message=question_message,
        question_text=question_text,
        source=source,
        question_type=question_type,
        difficulty=difficulty or session_metadata.get("difficulty", ""),
        rag_candidates=rag_candidates or [],
    )


def _answer_pending_turn(session_id: str, *, answer_message, answer_text: str):
    if not (STORAGE_AVAILABLE and data_client and turn_service):
        return None
    _ensure_turn_service_client()
    return turn_service.answer_pending_turn(
        session_id,
        answer_message=answer_message,
        answer_text=answer_text,
    )


def _record_agent_event(session_id: str, event_type: str, *, turn_id: str = "", payload: dict | None = None):
    if STORAGE_AVAILABLE and data_client and turn_service:
        _ensure_turn_service_client()
        turn_service.record_event(session_id, event_type, turn_id=turn_id, payload=payload or {})


def _skip_pending_turns_before_end(session_id: str) -> int:
    if not (STORAGE_AVAILABLE and data_client and hasattr(data_client, "skip_pending_turns")):
        return 0
    try:
        return int(data_client.skip_pending_turns(session_id) or 0)
    except Exception as exc:
        print(f"[app] skip pending turns failed: {exc}")
        return 0


def _retry_failed_turn_evaluations_before_end(session_id: str, observer, *, save_history: bool) -> int:
    if not observer or not save_history:
        return 0
    if not hasattr(observer, "retry_failed_turn_evaluations"):
        return 0
    try:
        return int(observer.retry_failed_turn_evaluations(session_id) or 0)
    except Exception as exc:
        print(f"[app] retry failed turn evaluations failed: {exc}")
        return 0


def _ensure_turn_service_client():
    if turn_service and getattr(turn_service, "data_client", None) is not data_client:
        turn_service.set_data_client(data_client)


def _turn_trace_metadata(turn: dict | None) -> dict:
    if not isinstance(turn, dict):
        return {}
    return {
        "turn_id": turn.get("turn_id"),
        "turn_no": turn.get("turn_no"),
    }


def _build_followup_quality_context(session_id: str, *, limit: int = 2) -> str:
    if not (STORAGE_AVAILABLE and data_client and hasattr(data_client, "list_turn_evaluations")):
        return ""

    try:
        rows = data_client.list_turn_evaluations(session_id) or []
    except Exception as exc:
        print(f"[app] build follow-up quality context failed: {exc}")
        return ""

    turn_status_by_id = _get_followup_turn_statuses(session_id)
    candidates = []
    for row in rows:
        status = turn_status_by_id.get(str(row.get("turn_id") or ""))
        if status in {"pending", "skipped", "evaluation_failed"}:
            continue
        suggestion = str(row.get("suggestion") or "").strip()
        evidence = str(row.get("evidence") or "").strip()
        if not suggestion and not evidence:
            continue
        try:
            score = int(row.get("score") or 0)
        except Exception:
            score = 0
        candidates.append({
            "turn_no": int(row.get("turn_no") or 0),
            "dimension": str(row.get("dimension") or "综合表现").strip() or "综合表现",
            "score": score,
            "pass_level": str(row.get("pass_level") or "").strip(),
            "suggestion": suggestion,
            "evidence": evidence,
        })

    if not candidates:
        return ""

    candidates.sort(key=lambda item: (item["turn_no"], bool(item["suggestion"])), reverse=True)
    lines = []
    seen = set()
    for item in candidates:
        key = (item["turn_no"], item["dimension"])
        if key in seen:
            continue
        seen.add(key)
        line = _summarize_followup_evaluation_item(item)
        if line:
            lines.append(line)
        if len(lines) >= limit:
            break

    if not lines:
        return ""

    return (
        "【隐藏面试官笔记】\n"
        "以下内容来自后台异步评估，只用于决定下一问方向，禁止直接复述，禁止向候选人提及评分、维度、rubric、evidence、suggestion 或系统元数据。\n"
        "请把它改写成自然、简短的一句追问；如果当前回答更适合推进新主题，可以忽略。\n"
        + "\n".join(lines)
    )


def _build_interviewer_hidden_context(session_id: str) -> str:
    sections = []
    compacted_context = _build_context_compaction_context(session_id)
    if compacted_context:
        sections.append(compacted_context)
    followup_context = _build_followup_quality_context(session_id)
    if followup_context:
        sections.append(followup_context)
    return "\n\n".join(sections)


def _estimate_text_tokens(value: str) -> int:
    return max(1, len(str(value or "")) // 4)


def _estimate_interview_context_tokens(session_id: str) -> int:
    if not (STORAGE_AVAILABLE and data_client):
        return 0
    try:
        turns = data_client.list_interview_turns(session_id) if hasattr(data_client, "list_interview_turns") else []
    except Exception:
        turns = []
    total_chars = 0
    for turn in turns or []:
        if not isinstance(turn, dict):
            continue
        total_chars += len(str(turn.get("question_text") or ""))
        total_chars += len(str(turn.get("answer_text") or ""))
    return _estimate_text_tokens("x" * total_chars) if total_chars else 0


def _build_context_compaction_context(session_id: str) -> str:
    checkpoint = _get_or_create_context_checkpoint(session_id)
    if not checkpoint:
        return ""
    lines = [
        "【隐藏长期记忆卡】",
        "以下内容来自结构化历史压缩，只用于避免重复提问、延续早期线索和识别未覆盖风险；不要向候选人展示或提及这张卡。",
    ]
    if checkpoint.get("recent_turns"):
        lines.append("最近关键轮次：" + "；".join(checkpoint["recent_turns"][:4]))
    if checkpoint.get("covered_dimensions"):
        lines.append("已覆盖能力：" + "、".join(checkpoint["covered_dimensions"][:8]))
    if checkpoint.get("candidate_facts"):
        lines.append("候选人已确认事实：" + "；".join(checkpoint["candidate_facts"][:5]))
    if checkpoint.get("risk_signals"):
        lines.append("风险信号：" + "；".join(checkpoint["risk_signals"][:5]))
    if checkpoint.get("open_threads"):
        lines.append("待追问线索：" + "；".join(checkpoint["open_threads"][:5]))
    return "\n".join(lines)


def _get_or_create_context_checkpoint(session_id: str) -> dict:
    if not (STORAGE_AVAILABLE and data_client):
        return {}
    try:
        turns = data_client.list_interview_turns(session_id) if hasattr(data_client, "list_interview_turns") else []
        metadata_rows = data_client.list_question_metadata(session_id) if hasattr(data_client, "list_question_metadata") else []
        evaluation_rows = data_client.list_turn_evaluations(session_id) if hasattr(data_client, "list_turn_evaluations") else []
    except Exception as exc:
        print(f"[app] build context checkpoint failed: {exc}")
        return {}

    answered_turns = [
        turn for turn in turns or []
        if isinstance(turn, dict) and str(turn.get("answer_text") or "").strip()
    ]
    if len(answered_turns) < CONTEXT_COMPACTION_MIN_TURNS:
        return {}

    estimated_tokens = _estimate_interview_context_tokens(session_id)
    threshold = int(CONTEXT_COMPACTION_BUDGET_TOKENS * CONTEXT_COMPACTION_TRIGGER_RATIO)
    if estimated_tokens < threshold:
        return {}

    last_turn_no = max(int(turn.get("turn_no") or 0) for turn in answered_turns)
    existing = _session_context_checkpoints.get(session_id) or {}
    if existing.get("last_turn_no") == last_turn_no:
        if existing.pop("_needs_memory_card_refresh", False):
            existing.update(_build_context_checkpoint_memory_fields(answered_turns, metadata_rows, evaluation_rows))
            existing["estimated_tokens"] = _positive_int(existing.get("estimated_tokens")) or estimated_tokens
            _session_context_checkpoints[session_id] = existing
        return existing
    if not existing:
        existing = _rehydrate_context_checkpoint_from_events(session_id, max_last_turn_no=last_turn_no)
        if existing.get("last_turn_no") == last_turn_no:
            if existing.pop("_needs_memory_card_refresh", False):
                existing.update(_build_context_checkpoint_memory_fields(answered_turns, metadata_rows, evaluation_rows))
                existing["estimated_tokens"] = _positive_int(existing.get("estimated_tokens")) or estimated_tokens
                _session_context_checkpoints[session_id] = existing
            return existing

    checkpoint = {
        "context_version": int(existing.get("context_version") or 1) + 1,
        "last_turn_no": last_turn_no,
        "estimated_tokens": estimated_tokens,
        **_build_context_checkpoint_memory_fields(answered_turns, metadata_rows, evaluation_rows),
    }
    _session_context_checkpoints[session_id] = checkpoint
    _record_agent_event(
        session_id,
        "context_compacted",
        payload=_context_checkpoint_event_payload(checkpoint, threshold),
    )
    trace_context = _session_trace_contexts.get(session_id)
    if isinstance(trace_context, dict):
        trace_context["context_version"] = checkpoint["context_version"]
    return checkpoint


def _build_context_checkpoint_memory_fields(answered_turns: list, metadata_rows: list, evaluation_rows: list) -> dict:
    metadata_by_turn = {item.get("turn_id"): item for item in metadata_rows if isinstance(item, dict)}
    evaluations_by_turn = {}
    for row in evaluation_rows or []:
        if isinstance(row, dict):
            evaluations_by_turn.setdefault(row.get("turn_id"), []).append(row)
    return {
        "recent_turns": _checkpoint_recent_turns(answered_turns),
        "covered_dimensions": _checkpoint_covered_dimensions(metadata_rows, evaluation_rows),
        "candidate_facts": _checkpoint_candidate_facts(answered_turns),
        "risk_signals": _checkpoint_risk_signals(answered_turns, evaluations_by_turn),
        "open_threads": _checkpoint_open_threads(answered_turns, metadata_by_turn, evaluations_by_turn),
    }


def _rehydrate_context_checkpoint_from_events(session_id: str, *, max_last_turn_no: int | None = None) -> dict:
    if not (STORAGE_AVAILABLE and data_client and hasattr(data_client, "list_agent_events")):
        return {}
    try:
        events = data_client.list_agent_events(session_id, event_type="context_compacted", limit=5) or []
    except Exception as exc:
        print(f"[app] rehydrate context checkpoint failed: {exc}")
        return {}

    for event in events:
        if not isinstance(event, dict):
            continue
        checkpoint = _context_checkpoint_from_event_payload(_agent_event_payload(event))
        if not checkpoint:
            continue
        if max_last_turn_no is not None and int(checkpoint.get("last_turn_no") or 0) > max_last_turn_no:
            continue
        _session_context_checkpoints[session_id] = checkpoint
        trace_context = _session_trace_contexts.get(session_id)
        if isinstance(trace_context, dict):
            trace_context["context_version"] = checkpoint["context_version"]
        return checkpoint
    return {}


def _agent_event_payload(event: dict) -> dict:
    payload = event.get("payload")
    if isinstance(payload, dict):
        return payload
    payload_json = event.get("payload_json")
    if isinstance(payload_json, str) and payload_json.strip():
        try:
            value = json_mod.loads(payload_json)
            return value if isinstance(value, dict) else {}
        except Exception:
            return {}
    return {}


def _context_checkpoint_from_event_payload(payload: dict) -> dict:
    if not isinstance(payload, dict):
        return {}
    last_turn_no = _positive_int(payload.get("last_turn_no"))
    if not last_turn_no:
        return {}
    memory_keys = ("recent_turns", "covered_dimensions", "candidate_facts", "risk_signals", "open_threads")
    return {
        "context_version": _positive_int(payload.get("context_version")) or 1,
        "last_turn_no": last_turn_no,
        "estimated_tokens": _positive_int(payload.get("estimated_tokens")),
        "recent_turns": _checkpoint_payload_list(payload.get("recent_turns"), limit=4),
        "covered_dimensions": _checkpoint_payload_list(payload.get("covered_dimensions"), limit=10),
        "candidate_facts": _checkpoint_payload_list(payload.get("candidate_facts"), limit=6),
        "risk_signals": _checkpoint_payload_list(payload.get("risk_signals"), limit=6),
        "open_threads": _checkpoint_payload_list(payload.get("open_threads"), limit=6),
        "_needs_memory_card_refresh": not all(key in payload for key in memory_keys),
    }


def _context_checkpoint_event_payload(checkpoint: dict, threshold_tokens: int) -> dict:
    return {
        "context_version": _positive_int(checkpoint.get("context_version")) or 1,
        "last_turn_no": _positive_int(checkpoint.get("last_turn_no")),
        "estimated_tokens": _positive_int(checkpoint.get("estimated_tokens")),
        "threshold_tokens": threshold_tokens,
        "recent_turns": _checkpoint_payload_list(checkpoint.get("recent_turns"), limit=4),
        "covered_dimensions": _checkpoint_payload_list(checkpoint.get("covered_dimensions"), limit=10),
        "candidate_facts": _checkpoint_payload_list(checkpoint.get("candidate_facts"), limit=6),
        "risk_signals": _checkpoint_payload_list(checkpoint.get("risk_signals"), limit=6),
        "open_threads": _checkpoint_payload_list(checkpoint.get("open_threads"), limit=6),
        "open_thread_count": len(checkpoint.get("open_threads") or []),
    }


def _checkpoint_payload_list(value, *, limit: int) -> list:
    if not isinstance(value, list):
        return []
    return _dedupe_nonempty(value)[:limit]


def _positive_int(value) -> int:
    try:
        number = int(value)
    except Exception:
        return 0
    return number if number > 0 else 0


def _checkpoint_recent_turns(turns: list) -> list:
    items = []
    for turn in turns[-4:]:
        question = _short_text(turn.get("question_text"), 48)
        answer = _short_text(turn.get("answer_text"), 72)
        if question or answer:
            items.append(f"第{turn.get('turn_no')}轮 问:{question} 答:{answer}")
    return items


def _checkpoint_covered_dimensions(metadata_rows: list, evaluation_rows: list) -> list:
    values = []
    for row in metadata_rows or []:
        if not isinstance(row, dict):
            continue
        for dimension in row.get("dimensions") or []:
            if isinstance(dimension, dict):
                values.append(str(dimension.get("name") or "").strip())
    for row in evaluation_rows or []:
        if isinstance(row, dict):
            values.append(str(row.get("dimension") or "").strip())
    return _dedupe_nonempty(values)[:10]


def _checkpoint_candidate_facts(turns: list) -> list:
    facts = []
    for turn in turns:
        answer = str(turn.get("answer_text") or "")
        snippets = re.split(r"[。！？!?；;\n]", answer)
        for snippet in snippets:
            text = _short_text(snippet, 90)
            if len(text) < 12:
                continue
            if any(keyword in text for keyword in ("负责", "做过", "使用", "优化", "设计", "参与", "实现", "项目")):
                facts.append(text)
                break
    return _dedupe_nonempty(facts)[-6:]


def _checkpoint_risk_signals(turns: list, evaluations_by_turn: dict) -> list:
    risks = []
    for turn in turns:
        answer = str(turn.get("answer_text") or "")
        if any(keyword in answer for keyword in ("不太清楚", "没有", "不确定", "大概", "应该")):
            risks.append(f"第{turn.get('turn_no')}轮回答存在不确定表达")
        for evaluation in evaluations_by_turn.get(turn.get("turn_id"), []):
            score = _coerce_report_score(evaluation.get("score"))
            if score and score < 7:
                dimension = evaluation.get("dimension") or "综合表现"
                evidence = _short_text(evaluation.get("evidence"), 80)
                risks.append(f"第{turn.get('turn_no')}轮{dimension}不够充分：{evidence}")
    return _dedupe_nonempty(risks)[-6:]


def _checkpoint_open_threads(turns: list, metadata_by_turn: dict, evaluations_by_turn: dict) -> list:
    threads = []
    for turn in turns:
        turn_no = turn.get("turn_no")
        for evaluation in evaluations_by_turn.get(turn.get("turn_id"), []):
            suggestion = _short_text(evaluation.get("suggestion"), 90)
            if suggestion:
                threads.append(f"第{turn_no}轮后续：{suggestion}")
        metadata = metadata_by_turn.get(turn.get("turn_id")) or {}
        for dimension in (metadata.get("dimensions") or [])[:1]:
            if isinstance(dimension, dict):
                pass_criteria = _short_text(dimension.get("pass_criteria"), 80)
                if pass_criteria:
                    threads.append(f"确认第{turn_no}轮是否满足：{pass_criteria}")
    return _dedupe_nonempty(threads)[-6:]


def _dedupe_nonempty(values: list) -> list:
    result = []
    seen = set()
    for value in values or []:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _short_text(value, limit: int) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def _get_followup_turn_statuses(session_id: str) -> dict:
    if not hasattr(data_client, "list_interview_turns"):
        return {}
    try:
        turns = data_client.list_interview_turns(session_id) or []
    except Exception:
        return {}
    statuses = {}
    for turn in turns:
        if not isinstance(turn, dict):
            continue
        turn_id = str(turn.get("turn_id") or "")
        if turn_id:
            statuses[turn_id] = str(turn.get("status") or "").strip()
    return statuses


def _summarize_followup_evaluation_item(item: dict) -> str:
    dimension = _plain_dimension_name(item.get("dimension"))
    score = int(item.get("score") or 0)
    pass_level = str(item.get("pass_level") or "").strip()
    suggestion = _strip_internal_terms(str(item.get("suggestion") or ""))
    evidence = _strip_internal_terms(str(item.get("evidence") or ""))

    if score and score < 7:
        priority = "优先补问"
    elif pass_level in {"fail", "weak_pass"}:
        priority = "适合补问"
    else:
        priority = "可自然延展"

    focus = _derive_followup_focus(suggestion, evidence, dimension)
    if not focus:
        return ""

    turn_no = int(item.get("turn_no") or 0)
    prefix = f"- 第{turn_no}轮" if turn_no > 0 else "-"
    return f"{prefix}{dimension}表现{_score_hint(score, pass_level)}，{priority}：{focus}"


def _derive_followup_focus(suggestion: str, evidence: str, dimension: str) -> str:
    text = f"{suggestion}。{evidence}"
    keyword_map = [
        (("指标", "量化", "P95", "延迟", "耗时", "提升", "下降", "结果"), "要求候选人补充前后对比指标和结果数据"),
        (("根因", "定位", "排查", "瓶颈", "原因"), "追问问题定位过程、根因判断依据和排查顺序"),
        (("一致性", "并发", "降级", "高并发", "缓存一致"), "追问高并发下的一致性、降级或取舍方案"),
        (("取舍", "权衡", "方案", "替代"), "追问方案取舍、替代方案和为什么这样选"),
        (("负责", "贡献", "边界", "角色"), "澄清个人贡献边界和具体负责部分"),
        (("验证", "测试", "上线", "监控", "回滚"), "追问验证方式、上线监控和风险控制"),
    ]
    for keywords, focus in keyword_map:
        if any(keyword in text for keyword in keywords):
            return focus

    dimension = dimension or "综合表现"
    if suggestion:
        return f"围绕{dimension}补充一个更具体的细节问题"
    if evidence:
        return f"围绕{dimension}追问尚未展开的事实和判断依据"
    return ""


def _plain_dimension_name(value: str) -> str:
    text = str(value or "综合表现").strip()
    return text if text else "综合表现"


def _score_hint(score: int, pass_level: str) -> str:
    if score:
        if score < 5:
            return "偏弱"
        if score < 7:
            return "不够充分"
        if score >= 9:
            return "较强"
    if pass_level == "fail":
        return "偏弱"
    if pass_level == "weak_pass":
        return "不够充分"
    return ""


def _strip_internal_terms(text: str) -> str:
    cleaned = str(text or "")
    for token in (
        "suggestion",
        "evidence",
        "rubric",
        "dimension",
        "pass_criteria",
        "score",
        "评分",
        "维度",
        "证据",
        "建议",
    ):
        cleaned = cleaned.replace(token, "")
    return cleaned[:260]


def _is_assistant_error_response(text: str) -> bool:
    return str(text or "").strip().startswith("抱歉，系统遇到了问题")


def _build_debug_info(agent, session_id: str = ""):
    """收集 agent 调试信息"""
    if not agent:
        return {}
    provider_key = _session_models.get(session_id, "unknown")
    provider = get_provider(provider_key)
    info = {
        "system_prompt": getattr(agent, 'prompt', ''),
        "chat_history": agent.get_chat_history() if hasattr(agent, 'get_chat_history') else [],
        "agent_mode": "LangChain" if hasattr(agent, 'agent_executor') and agent.agent_executor else "Fallback",
        "tools_available": [tool.name for tool in agent.tools] if hasattr(agent, 'tools') else [],
        "model_provider": provider_key,
        "model_name": provider.model if provider else getattr(agent, 'model_name', 'unknown'),
        "model_label": provider.label if provider else provider_key,
        "data_service": {
            **STORAGE_STATUS,
        },
    }
    return info


def _build_rag_debug_details(
    *,
    query: str = "",
    job_title: str = "",
    difficulty: str = "",
    interview_type: str = "",
    style: str = "",
    stage: str = "",
    status: str = "empty",
    error: str = "",
    jobs: list | None = None,
    questions: list | None = None,
    scripts: list | None = None,
):
    jobs = jobs or []
    questions = questions or []
    scripts = scripts or []

    def _truncate(value: str, limit: int = 240) -> str:
        text = (value or "").strip()
        if len(text) <= limit:
            return text
        return text[:limit] + "..."

    return {
        "query": query,
        "job_title": job_title,
        "difficulty": difficulty,
        "interview_type": interview_type,
        "style": style,
        "stage": stage,
        "status": status,
        "error": error,
        "counts": {
            "jobs": len(jobs),
            "questions": len(questions),
            "scripts": len(scripts),
        },
        "jobs": [
            {
                "id": item.get("id"),
                "document": _truncate(item.get("document") or item.get("content") or "", 400),
                "metadata": item.get("metadata") or {},
            }
            for item in jobs
        ],
        "questions": [
            {
                "id": item.get("id"),
                "document": _truncate(item.get("document") or item.get("content") or "", 260),
                "metadata": item.get("metadata") or {},
            }
            for item in questions
        ],
        "scripts": [
            {
                "id": item.get("id"),
                "document": _truncate(item.get("document") or item.get("content") or "", 260),
                "metadata": item.get("metadata") or {},
            }
            for item in scripts
        ],
    }


def _build_resume_summary_fallback(resume_text: str, max_items: int = 8) -> str:
    """Use extracted resume lines as a deterministic fallback when LLM summary is empty."""
    text = str(resume_text or "").strip()
    if not text:
        return ""

    lines = []
    seen = set()
    for raw_line in text.replace("\r", "\n").split("\n"):
        line = re.sub(r"\s+", " ", raw_line).strip(" \t-•*|")
        if len(line) < 2:
            continue
        if line.startswith("以下是提取的内容"):
            continue
        if line in seen:
            continue
        seen.add(line)
        lines.append(line[:140])
        if len(lines) >= max_items:
            break

    if not lines:
        compact = re.sub(r"\s+", " ", text)
        return compact[:800]

    return "以下为基于简历原文抽取的候选人要点：\n" + "\n".join(f"- {line}" for line in lines)


def _build_job_title_candidates(job_title: str) -> list[str]:
    """Relax title matching so RAG can still hit entries like 前端工程师 / 后端工程师."""
    raw = normalize_job_title(job_title)
    candidates: list[str] = []

    def _add(value: str):
        text = re.sub(r"\s+", " ", str(value or "")).strip()
        if text and text not in candidates:
            candidates.append(text)

    _add(raw)

    relaxed = re.sub(r"^(高级|资深|中级|初级|高阶|专家级?)", "", raw or "").strip()
    relaxed = re.sub(r"(高级|资深|中级|初级)$", "", relaxed).strip()
    _add(relaxed)

    replacements = {
        "前端开发工程师": "前端工程师",
        "后端开发工程师": "后端工程师",
        "Java开发工程师": "Java工程师",
        "Python开发工程师": "Python工程师",
        "Golang开发工程师": "Golang工程师",
    }
    for source, target in replacements.items():
        if source in raw:
            _add(raw.replace(source, target))
        if relaxed and source in relaxed:
            _add(relaxed.replace(source, target))

    return candidates


def _runtime_config_allowed() -> bool:
    if os.getenv("PROVIEW_ALLOW_RUNTIME_CONFIG") == "1" or os.getenv("PROVIEW_DESKTOP_MODE") == "1":
        return True

    remote_addr = str(request.remote_addr or "").strip().lower()
    host = str(request.host or "").split(":", 1)[0].strip().lower()
    return (
        remote_addr in {"127.0.0.1", "::1", "::ffff:127.0.0.1"}
        or host in {"127.0.0.1", "localhost"}
    )


def _forbid_runtime_config_if_needed():
    if not _runtime_config_allowed():
        return jsonify({
            "status": "error",
            "message": "当前环境未开启运行时配置写入能力"
        }), 403
    return None


def _retrieve_rag_context(
    *,
    job_title: str,
    difficulty: str,
    interview_type: str,
    style: str,
    resume_text: str,
    stage: str,
):
    if not (STORAGE_AVAILABLE and data_client):
        return "", _build_rag_debug_details(
            query="",
            job_title=job_title,
            difficulty=difficulty,
            interview_type=interview_type,
            style=style,
            stage=stage,
            status="not_started",
        )

    resume_keywords = str(resume_text or "").strip()[:300]
    last_debug = _build_rag_debug_details(
        query="",
        job_title=job_title,
        difficulty=difficulty,
        interview_type=interview_type,
        style=style,
        stage=stage,
        status="empty",
    )

    for title_candidate in _build_job_title_candidates(job_title):
        rag_parts = []
        jobs = []
        questions = []
        scripts = []
        enriched_query = f"{title_candidate} {resume_keywords}".strip() or title_candidate

        try:
            jobs = data_client.search_job_descriptions(
                title_candidate,
                top_k=1,
                difficulty=difficulty,
                interview_type=interview_type,
            )
            if jobs:
                job = jobs[0]
                meta = job.get("metadata", {})
                must_have = ", ".join(meta.get("must_have_skills") or [])
                tech_tags = ", ".join(meta.get("tech_tags") or []) or meta.get("tags", "")
                rag_parts.append(
                    f"### 岗位画像\n"
                    f"- 岗位：{meta.get('job_name', '') or meta.get('canonical_job_title', '')}\n"
                    f"- 关键技能：{must_have}\n"
                    f"- 技术标签：{tech_tags}\n"
                    f"- 要求：{job.get('document', '')[:500]}"
                )

            questions = data_client.search_questions(
                enriched_query,
                job_filter=title_candidate,
                top_k=5,
                difficulty=difficulty,
                interview_type=interview_type,
                style=style,
                stage="core",
            )
            if questions:
                q_lines = []
                for idx, item in enumerate(questions, 1):
                    meta = item.get("metadata", {})
                    q_lines.append(
                        f"{idx}. [{meta.get('dimension', '')}] {(item.get('document', '') or item.get('content', ''))[:200]}"
                    )
                    if meta.get("score_5"):
                        q_lines.append(f"   5分标准：{meta['score_5'][:150]}")
                    if meta.get("score_1"):
                        q_lines.append(f"   1分标准：{meta['score_1'][:100]}")
                rag_parts.append("### 推荐面试题及评分标准\n" + "\n".join(q_lines))

            scripts = data_client.search_hr_scripts(
                f"{title_candidate} {style} 面试 开场 自我介绍",
                stage="开场",
                top_k=2,
                interview_type=interview_type,
                style=style,
            )
            if scripts:
                s_lines = [
                    f"- [{item.get('metadata', {}).get('stage', '')}] {(item.get('document', '') or item.get('content', ''))[:200]}"
                    for item in scripts
                ]
                rag_parts.append("### 参考话术\n" + "\n".join(s_lines))

            debug = _build_rag_debug_details(
                query=enriched_query,
                job_title=title_candidate,
                difficulty=difficulty,
                interview_type=interview_type,
                style=style,
                stage=stage,
                status="matched" if rag_parts else "empty",
                jobs=jobs,
                questions=questions,
                scripts=scripts,
            )
            last_debug = debug

            if rag_parts:
                return "\n\n".join(rag_parts), debug
        except Exception as exc:
            last_debug = _build_rag_debug_details(
                query=enriched_query,
                job_title=title_candidate,
                difficulty=difficulty,
                interview_type=interview_type,
                style=style,
                stage=stage,
                status="error",
                error=str(exc),
                jobs=jobs,
                questions=questions,
                scripts=scripts,
            )
            print(f"[WARN] RAG retrieval failed for {title_candidate}: {exc}")

    return "", last_debug


def _resolve_session_owner_id(session_id: str):
    if not (STORAGE_AVAILABLE and data_client):
        return None

    try:
        session_info = data_client.get_session_info(session_id) or {}
    except Exception:
        return None

    return session_info.get("user_id")


def _parse_save_history_payload(default: bool = True) -> bool:
    data = request.get_json(silent=True) or {}
    raw = data.get("save_history", default)
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, str):
        return raw.strip().lower() not in {"false", "0", "no", "off"}
    return bool(raw)


def _has_self_intro_request(response_text: str) -> bool:
    text = (response_text or "").strip()
    if not text:
        return False
    keywords = [
        "自我介绍",
        "介绍一下自己",
        "介绍一下",
        "介绍自己",
        "简要介绍自己",
        "简单介绍自己",
        "先做一个介绍",
        "先做个介绍",
    ]
    if any(keyword in text for keyword in keywords):
        return True
    if ("2-3分钟" in text or "两三分钟" in text or "几分钟" in text) and "介绍" in text:
        return True
    return False


_STAGE_DIRECTION_BRACKETS = {
    "（": "）",
    "(": ")",
    "【": "】",
    "[": "]",
}
_STAGE_DIRECTION_PREFIXES = (
    "语气",
    "口吻",
    "语调",
    "语速",
    "停顿",
    "动作",
    "神态",
    "表情",
    "情绪",
    "态度",
)
_STAGE_DIRECTION_KEYWORDS = (
    "语气",
    "口吻",
    "语调",
    "语速",
    "停顿",
    "沉默",
    "冷静",
    "平静",
    "严肃",
    "温和",
    "柔和",
    "微笑",
    "轻笑",
    "低声",
    "轻声",
    "压低声音",
    "提高语速",
    "放慢语速",
    "先别寒暄",
    "切入正题",
    "开门见山",
    "不必重复问候",
    "不要重复问候",
    "不必问候",
    "不要问候",
    "直接开始",
    "带压迫感",
    "略带",
)
_STAGE_DIRECTION_KEYWORDS_EN = (
    "pause",
    "calm tone",
    "serious tone",
    "gentle tone",
    "softly",
    "slow down",
    "speed up",
    "skip greeting",
    "no greeting",
)


def _looks_like_stage_direction(segment: str) -> bool:
    inner = str(segment or "")[1:-1].strip()
    if not inner:
        return False

    compact = re.sub(r"\s+", "", inner)
    if not compact or len(compact) > 48:
        return False

    lower_inner = inner.lower()
    if any(keyword in lower_inner for keyword in _STAGE_DIRECTION_KEYWORDS_EN):
        return True

    if re.search(r"[A-Za-z0-9_/+#-]{2,}", compact):
        return False

    if compact in {"冷静", "平静", "严肃", "温和", "柔和", "停顿", "微笑", "轻笑"}:
        return True

    if any(compact.startswith(prefix) for prefix in _STAGE_DIRECTION_PREFIXES):
        return True

    if any(keyword in compact for keyword in _STAGE_DIRECTION_KEYWORDS):
        return True

    if re.fullmatch(r"(请)?(直接|不要|不必).{0,18}(问候|寒暄|进入正题|开始)", compact):
        return True

    if re.fullmatch(r".{0,12}(地说|地问|地回应|地追问|地开场)", compact):
        return True

    return False


class _StreamStageDirectionSanitizer:
    def __init__(self):
        self._candidate_chars = []
        self._candidate_close = ""
        self._drop_prefix = False

    def feed(self, text: str) -> str:
        if not text:
            return ""

        output = []
        for ch in text:
            if self._drop_prefix and ch in " \t\r\n\u3000:：,，-—":
                continue
            self._drop_prefix = False

            if not self._candidate_chars:
                close = _STAGE_DIRECTION_BRACKETS.get(ch)
                if close:
                    self._candidate_chars = [ch]
                    self._candidate_close = close
                else:
                    output.append(ch)
                continue

            self._candidate_chars.append(ch)
            if ch == self._candidate_close:
                segment = "".join(self._candidate_chars)
                if _looks_like_stage_direction(segment):
                    self._drop_prefix = True
                else:
                    output.append(segment)
                self._candidate_chars = []
                self._candidate_close = ""
            elif len(self._candidate_chars) > 64:
                output.append("".join(self._candidate_chars))
                self._candidate_chars = []
                self._candidate_close = ""

        return "".join(output)

    def flush(self) -> str:
        if not self._candidate_chars:
            return ""
        segment = "".join(self._candidate_chars)
        self._candidate_chars = []
        self._candidate_close = ""
        return segment


def _sanitize_spoken_text(text: str) -> str:
    if not text:
        return ""

    sanitizer = _StreamStageDirectionSanitizer()
    cleaned = sanitizer.feed(text) + sanitizer.flush()
    cleaned = re.sub(r"[ \t]{2,}", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _sync_last_assistant_message(agent, content: str) -> None:
    if not agent:
        return

    history = getattr(agent, "chat_history", None)
    if not isinstance(history, list):
        return

    for idx in range(len(history) - 1, -1, -1):
        item = history[idx]
        if isinstance(item, dict) and item.get("role") == "assistant":
            item["content"] = content
            return


def _sanitize_assistant_response(text: str, agent=None) -> str:
    cleaned = _sanitize_spoken_text(text)
    if agent:
        _sync_last_assistant_message(agent, cleaned)
    return cleaned


# PromptGenerator 延迟导入（问题4完整实现后启用）
_prompt_generator = None
EVAL_OBSERVER_DRAIN_TIMEOUT_SECONDS = 3.0


def _try_generate_prompt(job_title, interview_type, difficulty, style, resume_summary, job_requirements=""):
    """尝试使用 PromptGenerator 生成定制 prompt，失败返回空字符串（降级到静态模板）"""
    global _prompt_generator
    try:
        if _prompt_generator is None:
            from core.prompt_generator import PromptGenerator
            _prompt_generator = PromptGenerator(
                api_key=DEEPSEEK_API_KEY,
                base_url=DEEPSEEK_BASE_URL
            )
        return _prompt_generator.generate(
            job_title=job_title,
            interview_type=interview_type,
            difficulty=difficulty,
            style=style,
            resume_summary=resume_summary,
            job_requirements=job_requirements,
        )
    except ImportError:
        # prompt_generator.py 尚未创建，静默降级
        return ""
    except Exception as e:
        print(f"[WARN] PromptGenerator call failed, fallback to static template: {e}")
        return ""


@app.route('/api/health', methods=['GET'])
def health():
    _refresh_storage_status()
    return jsonify({
        "status": "ok",
        "data_service": {
            **STORAGE_STATUS,
        },
        "agent_available": AGENT_AVAILABLE,
        "ocr_available": OCR_AVAILABLE,
        "runtime_config_enabled": _runtime_config_allowed(),
    })


@app.route('/api/runtime-config', methods=['GET'])
def get_runtime_config():
    forbidden = _forbid_runtime_config_if_needed()
    if forbidden:
        return forbidden

    snapshot = app_config.get_runtime_config_snapshot(mask_secrets=True)
    return jsonify({
        "status": "success",
        **snapshot,
        "models": list_available_providers(),
        "speech_available": SPEECH_AVAILABLE,
    })


@app.route('/api/runtime-config', methods=['POST'])
def update_runtime_config():
    forbidden = _forbid_runtime_config_if_needed()
    if forbidden:
        return forbidden

    payload = request.get_json(silent=True) or {}
    fields = payload.get("fields") if isinstance(payload, dict) else None
    if not isinstance(fields, dict):
        return jsonify({
            "status": "error",
            "message": "fields 必须是对象"
        }), 400

    updates = {
        key: value
        for key, value in fields.items()
        if key in app_config.RUNTIME_CONFIG_FIELDS
    }
    if not updates:
        return jsonify({
            "status": "error",
            "message": "未提供可更新的配置项"
        }), 400

    snapshot = app_config.persist_runtime_config(updates)
    _reload_runtime_config_state()
    _refresh_speech_client()

    return jsonify({
        "status": "success",
        **snapshot,
        "models": list_available_providers(),
        "speech_available": SPEECH_AVAILABLE,
        "message": "运行配置已保存",
    })


@app.route('/api/models', methods=['GET'])
def get_models():
    """返回可用的模型提供商列表"""
    return jsonify({"models": list_available_providers()})


@app.route('/api/positions', methods=['GET'])
def get_positions():
    """返回岗位列表"""
    return jsonify({"positions": _positions})


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    if not STORAGE_AVAILABLE:
        return jsonify({"status": "error", "message": "数据服务不可用"}), 503
    user = _get_current_user_info()
    if not user:
        return jsonify({"status": "error", "message": "本机用户初始化失败"}), 500
    return jsonify({"status": "success", "user": user})


# ── 面试历史 ──

@app.route('/api/history/sessions', methods=['GET'])
def list_user_sessions():
    """获取当前本机用户的面试历史列表"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify([])
    user_info = _get_current_user_info()
    if not user_info or "id" not in user_info:
        return jsonify([])
    user_id = user_info["id"]
    sessions = data_client.list_sessions(limit=None, user_id=user_id)
    # 兼容旧数据：如果按 user_id 查不到，也返回 user_id 为空的 session
    if not sessions and not _is_new_user(user_info):
        all_sessions = data_client.list_sessions(limit=None)
        sessions = [s for s in all_sessions if not s.get("user_id")]
    return jsonify(normalize_session_list(sessions))


@app.route('/api/history/sessions/<session_id>', methods=['GET'])
def get_session_detail(session_id):
    """获取某次面试的完整详情（元信息 + 聊天记录 + 评分）"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"error": "数据服务不可用"}), 503
    session_info = normalize_session_info(data_client.get_session_info(session_id))
    if not session_info:
        return jsonify({"error": "会话不存在"}), 404
    messages = data_client.get_session_history(session_id)
    stats = data_client.get_session_statistics(session_id)
    return jsonify({
        "session": session_info,
        "messages": messages,
        "stats": stats,
    })


@app.route('/api/history/sessions/<session_id>', methods=['DELETE'])
def delete_session_detail(session_id):
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"status": "error", "message": "storage unavailable"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    deleted = data_client.delete_session(session_id, user_id)
    if not deleted:
        return jsonify({"status": "error", "message": "session not found"}), 404

    return jsonify({
        "status": "success",
        "quota": _build_history_quota(user_id),
    })


@app.route('/api/history/quota', methods=['GET'])
def get_history_quota():
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"status": "error", "message": "storage unavailable"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    return jsonify(_build_history_quota(user_id))


@app.route('/api/history/resume/<session_id>', methods=['GET'])
def get_session_resume(session_id):
    """获取某次面试关联的简历 OCR 文本"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify(None)
    resume = data_client.get_resume_by_session(session_id)
    return jsonify(resume)


@app.route('/api/history/resume/latest', methods=['GET'])
def get_latest_resume():
    """获取当前用户最近一条有 OCR 结果的简历"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify(None)
    user_id = _get_current_user_id_from_auth_header()
    resume = data_client.get_latest_resume(user_id=user_id)
    return jsonify(resume)


@app.route('/api/my-resumes', methods=['GET'])
def list_my_resumes():
    """获取当前用户的所有简历列表"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify([])
    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify([])
    resumes = data_client.list_user_resumes(user_id)
    return jsonify([_serialize_resume_library_record(record) for record in resumes])


@app.route('/api/my-resumes/<int:resume_id>/file', methods=['GET'])
def get_resume_file(resume_id):
    """获取简历原始文件"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"error": "存储服务不可用"}), 503
    try:
        current_user_id = _get_current_user_id_from_auth_header()
        record = data_client.get_resume_file_record(resume_id, user_id=current_user_id)
        if not record:
            return jsonify({"error": "文件不存在"}), 404

        file_path = record.get("file_path") or ""
        if not file_path or not os.path.exists(file_path):
            return jsonify({"error": "文件不存在"}), 404

        return send_file(
            file_path,
            mimetype="application/octet-stream",
            as_attachment=False,
            download_name=record.get("file_name") or os.path.basename(file_path),
        )
    except Exception as e:
        print(f"[API] get_resume_file failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/my-resumes/<int:resume_id>', methods=['DELETE'])
def delete_my_resume(resume_id):
    """删除当前用户的简历，并清理服务器文件与预览图。"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"status": "error", "message": "存储服务不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    deleted = data_client.delete_resume(resume_id, user_id)
    if not deleted:
        return jsonify({"status": "error", "message": "简历不存在或无权删除"}), 404
    return jsonify({"status": "success"})


@app.route('/api/my-resumes/<int:resume_id>/preview/<int:page>', methods=['GET'])
def get_resume_preview(resume_id, page):
    """返回简历的图片化预览页。"""
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"error": "存储服务不可用"}), 503

    try:
        current_user_id = _get_current_user_id_from_auth_header()
        record = data_client.get_resume_file_record(resume_id, user_id=current_user_id)
        if not record:
            return jsonify({"error": "简历不存在"}), 404

        preview_summary = get_resume_preview_summary(record.get("file_path") or "", record.get("file_name") or "")
        preview_paths = preview_summary["preview_paths"]
        if page < 1 or page > len(preview_paths):
            return jsonify({"error": "预览页不存在"}), 404

        preview_path = preview_paths[page - 1]
        if not preview_path or not os.path.exists(preview_path):
            return jsonify({"error": "预览文件不存在"}), 404

        return send_file(preview_path, mimetype="image/png", as_attachment=False)
    except Exception as e:
        print(f"[API] get_resume_preview failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/career/dashboard', methods=['GET'])
def get_career_dashboard():
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    try:
        dashboard = career_planning_service.build_dashboard(user_id)
        return jsonify({"status": "success", "data": dashboard})
    except Exception as e:
        print(f"[API] get_career_dashboard failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/docs', methods=['GET'])
def get_career_docs():
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"status": "error", "message": "数据服务不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    try:
        return jsonify({"status": "success", "data": career_planning_docs.get_catalog()})
    except Exception as e:
        print(f"[API] get_career_docs failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/docs/<doc_id>', methods=['GET'])
def get_career_doc(doc_id):
    if not STORAGE_AVAILABLE or not data_client:
        return jsonify({"status": "error", "message": "数据服务不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    try:
        document = career_planning_docs.get_document(doc_id)
        if not document:
            return jsonify({"status": "error", "message": "文档不存在"}), 404
        return jsonify({"status": "success", "data": document})
    except Exception as e:
        print(f"[API] get_career_doc failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/plans', methods=['GET'])
def list_career_plans():
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    try:
        dashboard = career_planning_service.build_dashboard(user_id)
        return jsonify({"status": "success", "data": {"plans": dashboard.get("plans", [])}})
    except Exception as e:
        print(f"[API] list_career_plans failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/plans/generate', methods=['POST'])
def generate_career_plan():
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    payload = request.get_json(silent=True) or {}
    try:
        target_role, career_goal, horizon_months, refresh = _parse_career_plan_generate_payload(payload)
        dashboard = career_planning_service.generate_plan(
            user_id=user_id,
            target_role=target_role,
            career_goal=career_goal,
            horizon_months=horizon_months,
            refresh=refresh,
        )
        return jsonify({"status": "success", "data": dashboard})
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        print(f"[API] generate_career_plan failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/plans/<int:plan_id>', methods=['GET'])
def get_career_plan(plan_id):
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    try:
        detail = career_planning_service._fetch_plan_detail(plan_id, user_id)
        if not detail:
            return jsonify({"status": "error", "message": "规划不存在或无权限访问"}), 404
        return jsonify({"status": "success", "data": detail})
    except Exception as e:
        print(f"[API] get_career_plan failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/tasks/<int:task_id>', methods=['PATCH'])
def update_career_task(task_id):
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    payload = request.get_json(silent=True) or {}
    try:
        detail = career_planning_service.update_task(
            user_id=user_id,
            task_id=task_id,
            status=str(payload.get("status") or "").strip(),
            progress=payload.get("progress"),
            note=str(payload.get("note") or "").strip(),
        )
        if not detail:
            return jsonify({"status": "error", "message": "任务不存在或无权限访问"}), 404
        return jsonify({"status": "success", "data": career_planning_service.build_dashboard(user_id)})
    except Exception as e:
        print(f"[API] update_career_task failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/career/tasks/<int:task_id>/logs', methods=['POST'])
def append_career_task_log(task_id):
    if not STORAGE_AVAILABLE or not data_client or not career_planning_service:
        return jsonify({"status": "error", "message": "职业规划服务暂不可用"}), 503

    user_id = _get_current_user_id_from_auth_header()
    if user_id is None:
        return jsonify({"status": "error", "message": "本机用户不可用"}), 401

    payload = request.get_json(silent=True) or {}
    try:
        detail = career_planning_service.update_task(
            user_id=user_id,
            task_id=task_id,
            status=str(payload.get("status") or "").strip(),
            progress=payload.get("progress"),
            note=str(payload.get("note") or "").strip(),
        )
        if not detail:
            return jsonify({"status": "error", "message": "任务不存在或无权限访问"}), 404
        return jsonify({"status": "success", "data": career_planning_service.build_dashboard(user_id)})
    except Exception as e:
        print(f"[API] append_career_task_log failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/setup', methods=['POST'])
def setup_interview():
    """初始化面试配置与简历上传"""
    job_title = normalize_job_title(request.form.get('job_title', '高级前端开发工程师'))
    style = request.form.get('style', 'strict')
    interview_type = request.form.get('interview_type', 'technical')
    difficulty = request.form.get('difficulty', 'mid')
    feature_vad = request.form.get('feature_vad', 'true') == 'true'
    feature_deep = request.form.get('feature_deep', 'true') == 'true'
    model_provider = request.form.get('model_provider', '')
    resume_ocr_text = request.form.get('resume_ocr_text', '')
    job_requirements = request.form.get('job_requirements', '').strip()
    file_path = ""
    filename = ""
    explicit_resume_requested = False
    prepared_resume = None

    current_user_id = _get_current_user_id_from_auth_header()

    if 'resume' in request.files:
        file = request.files['resume']
        if file.filename != '':
            explicit_resume_requested = True
            try:
                filename, file_path = _save_uploaded_resume(file)
                prepared_resume = _extract_resume_payload(file_path)
                if not prepared_resume["success"]:
                    message = prepared_resume["error_message"] or "简历内容提取失败，请重新上传。"
                    _cleanup_local_resume_file(file_path)
                    return jsonify({"status": "error", "message": message}), 400
            except ResumeOcrUnavailableError as exc:
                _cleanup_local_resume_file(file_path)
                return jsonify({"status": "error", "message": str(exc)}), 503
            except ResumeExtractionError as exc:
                _cleanup_local_resume_file(file_path)
                return jsonify({"status": "error", "message": str(exc)}), 400
    elif resume_ocr_text:
        explicit_resume_requested = True
        prepared_resume = _build_reused_resume_payload(resume_ocr_text, source_label="历史简历")
        if not prepared_resume["success"]:
            return jsonify({"status": "error", "message": prepared_resume["error_message"]}), 400

    session_id = str(uuid.uuid4())
    token = create_token(session_id)

    if STORAGE_AVAILABLE:
        data_client.create_session(
            session_id=session_id,
            candidate_name="求职者",
            position=job_title,
            interview_style=style,
            metadata={
                "type": interview_type,
                "diff": difficulty,
                "vad": feature_vad,
                "deep": feature_deep,
                "job_requirements": job_requirements,
            },
            user_id=current_user_id,
        )

    agent = get_agent(session_id, model_provider=model_provider)
    _register_session_trace_context(
        session_id=session_id,
        user_id=current_user_id,
        job_title=job_title,
        interview_type=interview_type,
        difficulty=difficulty,
        style=style,
        has_resume=bool(prepared_resume),
        feature_vad=feature_vad,
        feature_deep=feature_deep,
    )

    try:
        resume_summary = ""
        ocr_raw_text = ""
        ocr_reusable_text = ""
        ocr_status = "not_called"  # not_called | success | error | unavailable
        prompt_source = "static"
        has_resume = bool(prepared_resume)

        if prepared_resume:
            ocr_raw_text = prepared_resume["text"]
            ocr_reusable_text = prepared_resume["reusable_text"]
            ocr_status = "success"

        if agent:
            parse_result_text = ""
            all_steps = []

            # ── 阶段 1：获取 OCR 文本（串行，后续步骤依赖它） ──
            if prepared_resume:
                has_resume = True
                ocr_raw_text = prepared_resume["text"]
                ocr_reusable_text = prepared_resume["reusable_text"]
                ocr_status = "success"
                print(f"[RESUME] Parsed resume via {prepared_resume['mode']}: {file_path or 'history_resume'}")
                all_steps.append({
                    "tool": "resume_reuse" if prepared_resume["mode"] == "reused_text" else ("perform_ocr" if prepared_resume["mode"] == "ocr" else "extract_resume_text"),
                    "tool_input": file_path or "历史简历文本",
                    "log": "自动选择 OCR、直接文本提取或复用已保存简历内容",
                    "observation": (ocr_reusable_text or prepared_resume["error_message"])[:2000]
                })

                if file_path and STORAGE_AVAILABLE:
                    stored_path = file_path
                    upload_result = data_client.upload_resume_file(session_id, file_path)
                    if upload_result and upload_result.get("file_path"):
                        stored_path = upload_result["file_path"]
                        print(f"[OK] Resume file saved: {stored_path}")
                    else:
                        print("[WARN] Resume file save failed, metadata only")
                    data_client.save_resume(
                        session_id,
                        filename,
                        stored_path,
                        normalize_reusable_ocr_result(ocr_reusable_text),
                        user_id=current_user_id,
                    )

            # ── 阶段 2：LLM 简历摘要 + RAG 检索 并行执行 ──
            rag_context = ""
            rag_debug_details = _build_rag_debug_details(
                query="",
                job_title=job_title,
                difficulty=difficulty,
                interview_type=interview_type,
                style=style,
                stage="opening",
                status="not_started",
            )

            def _task_summary():
                """LLM 生成简历摘要"""
                if ocr_status != "success" or not ocr_raw_text:
                    return "", []
                q = f"请用简洁的中文总结以下简历的核心信息（姓名、技能、项目经历、教育背景），不要遗漏关键细节：\n\n{ocr_raw_text[:6000]}"
                resp, steps = agent.run(
                    q,
                    trace_context=_build_langfuse_trace_context(
                        session_id,
                        interaction_type="opening.resume_summary",
                    ),
                )
                if not str(resp or "").strip():
                    fallback = _build_resume_summary_fallback(ocr_raw_text)
                    if fallback:
                        steps = steps + [{
                            "tool": "resume_summary_fallback",
                            "tool_input": "基于简历原文的确定性摘要回退",
                            "log": "LLM 摘要为空，回退到原文抽取摘要",
                            "observation": fallback[:500],
                        }]
                        return fallback, steps
                return resp, steps

            def _task_rag():
                """RAG 检索（用 OCR 原文片段代替 LLM 摘要做查询 enrichment）。"""
                return _retrieve_rag_context(
                    job_title=job_title,
                    difficulty=difficulty,
                    interview_type=interview_type,
                    style=style,
                    resume_text=ocr_raw_text,
                    stage="opening",
                )

            # 并行执行 LLM 摘要 + RAG 检索
            with ThreadPoolExecutor(max_workers=2) as executor:
                future_summary = executor.submit(_task_summary)
                future_rag = executor.submit(_task_rag)

                resume_summary, summary_steps = future_summary.result()
                all_steps.extend(summary_steps)
                rag_context, rag_debug_details = future_rag.result()

            if not resume_summary and ocr_status == "success" and ocr_raw_text:
                resume_summary = _build_resume_summary_fallback(ocr_raw_text)
                if resume_summary:
                    all_steps.append({
                        "tool": "resume_summary_fallback",
                        "tool_input": "基于简历原文的确定性摘要回退",
                        "log": "并行摘要阶段无结果，回退到原文抽取摘要",
                        "observation": resume_summary[:500],
                    })

            if explicit_resume_requested and has_resume and not resume_summary:
                return jsonify({
                    "status": "error",
                    "message": "简历文本已提取，但未能生成有效摘要，请重新上传内容更完整的简历。",
                }), 422

            print(f"[OK] Parallel stage complete: summary {len(resume_summary)} chars, RAG {len(rag_context)} chars")

            # 第三步：用简历摘要 + RAG 上下文重新注入 prompt（反幻觉锚点）
            if hasattr(agent, 'update_dynamic_prompt'):
                # 尝试使用 PromptGenerator 生成定制 prompt
                generated_prompt = ""
                try:
                    generated_prompt = _try_generate_prompt(
                        job_title, interview_type, difficulty, style, resume_summary, job_requirements
                    )
                    if generated_prompt:
                        prompt_source = "prompt_generator"
                except Exception as e:
                    print(f"[WARN] PromptGenerator degraded: {e}")

                agent.update_dynamic_prompt(
                    job_title=job_title,
                    interview_type=interview_type,
                    difficulty=difficulty,
                    style=style,
                    feature_vad=feature_vad,
                    feature_deep=feature_deep,
                    resume_summary=resume_summary,
                    job_requirements=job_requirements,
                    custom_prompt=generated_prompt,
                    rag_context=rag_context
                )
            else:
                agent.update_system_prompt(style=style)
            agent.reset_memory()

            # 第三步：让面试官开场
            if has_resume:
                setup_query = (
                    f"我应聘的岗位是【{job_title}】。你刚才已经解析了我的真实简历，现在面试正式开始。"
                    "请以面试官的身份用一段话向我打招呼，简述对我简历的第一印象，然后明确要求候选人做 2-3 分钟的自我介绍。"
                    "直接输出候选人最终会听到的话术，不要附带任何括号中的语气说明、动作说明或舞台指令。"
                    "切记：绝对不要自己编造任何经历！"
                )
                response, steps2 = agent.run(
                    setup_query,
                    trace_context=_build_langfuse_trace_context(
                        session_id,
                        interaction_type="opening.first_question",
                    ),
                )
                response = _sanitize_assistant_response(response, agent)
                
                # [CHECK] Verify: Check if AI's first sentence includes self-introduction request
                if _has_self_intro_request(response):
                    print(f"[OK] Verification passed: AI requires candidate self-introduction")
                else:
                    print(f"[FAIL] Verification failed: AI did not require self-introduction")
                    print(f"[AI] AI response: {response[:200]}")
                
                parse_result_text = "[OK] Resume parsed successfully, interviewer has locked real project details. Please introduce yourself first."
                all_steps = all_steps + steps2
            else:
                parse_result_text = "No resume provided"
                setup_query = (
                    f"我应聘的岗位是【{job_title}】。我没有提供简历。"
                    "请以面试官的身份向我打招呼，并要求候选人先做自我介绍。"
                    "直接输出候选人最终会听到的话术，不要附带任何括号中的语气说明、动作说明或舞台指令。"
                )
                response, steps_setup = agent.run(
                    setup_query,
                    trace_context=_build_langfuse_trace_context(
                        session_id,
                        interaction_type="opening.first_question",
                    ),
                )
                response = _sanitize_assistant_response(response, agent)
                
                # [CHECK] Verify: Check if AI's first sentence includes self-introduction request
                if _has_self_intro_request(response):
                    print(f"[OK] Verification passed: AI requires candidate self-introduction")
                else:
                    print(f"[FAIL] Verification failed: AI did not require self-introduction")
                    print(f"[AI] AI response: {response[:200]}")
                
                all_steps = all_steps + steps_setup

            assistant_message = _append_visible_message(session_id, "assistant", response)
            _create_pending_turn(
                session_id,
                question_message=assistant_message,
                question_text=response,
                source="interviewer_llm",
                question_type="opening",
                difficulty=difficulty,
                rag_candidates=rag_debug_details.get("questions") if isinstance(rag_debug_details, dict) else [],
            )

            debug_info = _build_debug_info(agent, session_id)
            debug_info["intermediate_steps"] = all_steps
            debug_info["prompt_source"] = prompt_source
            debug_info["resume_summary"] = resume_summary[:2000] if resume_summary else ""
            debug_info["ocr_raw_text"] = ocr_raw_text
            debug_info["ocr_status"] = ocr_status
            debug_info["rag_context"] = rag_context[:1000] if rag_context else ""
            debug_info["rag_details"] = rag_debug_details

            return jsonify({
                "status": "success",
                "token": token,
                "session_id": session_id,
                "system_message": f"[OK] Interview room created (ID: {session_id[:8]})",
                "parse_result": parse_result_text,
                "ai_response": response,
                "ocr_text": ocr_reusable_text if ocr_status == "success" else "",
                "debug_info": debug_info
            })
        else:
            import time
            time.sleep(1)
            ai_resp = f"你好，我是今天的AI面试官。我已经了解了你应聘【{job_title}】的意向。"
            parse_result_mock = ""
            if has_resume:
                ai_resp += f"我也初步看了你的简历（{filename}），发现你在项目经历上有几个点很值得探讨。准备好的话，请先做一个结合项目的简单自我介绍吧。"
                parse_result_mock = "发现简历中存在跨度较大的项目经历，将作为重点深挖对象。"
            else:
                ai_resp += "准备好的话，请先做一个简单的自我介绍吧。"
            ai_resp = _sanitize_assistant_response(ai_resp)
            assistant_message = _append_visible_message(session_id, "assistant", ai_resp)
            _create_pending_turn(
                session_id,
                question_message=assistant_message,
                question_text=ai_resp,
                source="mock",
                question_type="opening",
                difficulty=difficulty,
                rag_candidates=[],
            )
            return jsonify({
                "status": "success",
                "token": token,
                "session_id": session_id,
                "system_message": "[Mock 模式] 未检测到 Agent，使用模拟逻辑",
                "parse_result": parse_result_mock,
                "ai_response": ai_resp,
                "ocr_text": ocr_reusable_text if ocr_status == "success" else "",
            })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/setup-stream', methods=['POST'])
def setup_interview_stream():
    """流式面试初始化：通过 SSE 实时输出 LLM 思考过程。"""
    job_title = normalize_job_title(request.form.get('job_title', '高级前端开发工程师'))
    style = request.form.get('style', 'strict')
    interview_type = request.form.get('interview_type', 'technical')
    difficulty = request.form.get('difficulty', 'mid')
    feature_vad = request.form.get('feature_vad', 'true') == 'true'
    feature_deep = request.form.get('feature_deep', 'true') == 'true'
    model_provider = request.form.get('model_provider', '')
    resume_ocr_text = request.form.get('resume_ocr_text', '')
    job_requirements = request.form.get('job_requirements', '').strip()
    file_path = ""
    filename = ""
    explicit_resume_requested = False
    prepared_resume = None

    current_user_id = _get_current_user_id_from_auth_header()

    if 'resume' in request.files:
        file = request.files['resume']
        if file.filename != '':
            explicit_resume_requested = True
            try:
                filename, file_path = _save_uploaded_resume(file)
                prepared_resume = _extract_resume_payload(file_path)
                if not prepared_resume["success"]:
                    message = prepared_resume["error_message"] or "简历内容提取失败，请重新上传。"
                    _cleanup_local_resume_file(file_path)
                    return jsonify({"status": "error", "message": message}), 400
            except ResumeOcrUnavailableError as exc:
                _cleanup_local_resume_file(file_path)
                return jsonify({"status": "error", "message": str(exc)}), 503
            except ResumeExtractionError as exc:
                _cleanup_local_resume_file(file_path)
                return jsonify({"status": "error", "message": str(exc)}), 400
    elif resume_ocr_text:
        explicit_resume_requested = True
        prepared_resume = _build_reused_resume_payload(resume_ocr_text, source_label="历史简历")
        if not prepared_resume["success"]:
            return jsonify({"status": "error", "message": prepared_resume["error_message"]}), 400

    session_id = str(uuid.uuid4())
    token = create_token(session_id)

    if STORAGE_AVAILABLE:
        data_client.create_session(
            session_id=session_id, candidate_name="求职者",
            position=job_title, interview_style=style,
            metadata={
                "type": interview_type,
                "diff": difficulty,
                "vad": feature_vad,
                "deep": feature_deep,
                "job_requirements": job_requirements,
            },
            user_id=current_user_id,
        )

    agent = get_agent(session_id, model_provider=model_provider)
    _register_session_trace_context(
        session_id=session_id,
        user_id=current_user_id,
        job_title=job_title,
        interview_type=interview_type,
        difficulty=difficulty,
        style=style,
        has_resume=bool(prepared_resume),
        feature_vad=feature_vad,
        feature_deep=feature_deep,
    )

    def generate():
        nonlocal file_path, filename, prepared_resume
        try:
            resume_summary = ""
            ocr_raw_text = ""
            ocr_reusable_text = ""
            ocr_status = "not_called"
            prompt_source = "static"
            all_steps = []
            has_resume = bool(prepared_resume)

            # 阶段 1：加载简历
            if prepared_resume:
                if file_path:
                    yield _sse_event("stage", {"stage": "正在解析简历..."})
                else:
                    yield _sse_event("stage", {"stage": "正在加载已保存简历..."})

                ocr_raw_text = prepared_resume["text"]
                ocr_reusable_text = prepared_resume["reusable_text"]
                ocr_status = "success"
                all_steps.append({
                    "tool": "resume_reuse" if prepared_resume["mode"] == "reused_text" else ("perform_ocr" if prepared_resume["mode"] == "ocr" else "extract_resume_text"),
                    "tool_input": file_path or "历史简历文本",
                    "log": "自动选择 OCR、直接文本提取或复用已保存简历内容",
                    "observation": (ocr_reusable_text or prepared_resume["error_message"])[:2000],
                })

                if file_path and STORAGE_AVAILABLE:
                    stored_path = file_path
                    upload_result = data_client.upload_resume_file(session_id, file_path)
                    if upload_result and upload_result.get("file_path"):
                        stored_path = upload_result["file_path"]
                    data_client.save_resume(
                        session_id,
                        filename,
                        stored_path,
                        normalize_reusable_ocr_result(ocr_reusable_text),
                        user_id=current_user_id,
                    )
            if not agent:
                import time
                time.sleep(1)
                ai_resp = f"你好，我是今天的AI面试官。我已经了解了你应聘【{job_title}】的意向。"
                parse_result_text = ""
                if has_resume:
                    ai_resp += "我已经看过你的简历，准备好的话，请先结合你的经历做一个简单的自我介绍吧。"
                    parse_result_text = "[OK] Resume parsed successfully, interviewer has locked real project details. Please introduce yourself first."
                else:
                    ai_resp += "准备好的话，请先做一个简单的自我介绍吧。"
                ai_resp = _sanitize_assistant_response(ai_resp)
                assistant_message = _append_visible_message(session_id, "assistant", ai_resp)
                _create_pending_turn(
                    session_id,
                    question_message=assistant_message,
                    question_text=ai_resp,
                    source="mock",
                    question_type="opening",
                    difficulty=difficulty,
                    rag_candidates=[],
                )
                yield _sse_event("done", {
                    "status": "success", "token": token,
                    "session_id": session_id,
                    "system_message": "[Mock 模式]", "parse_result": parse_result_text,
                    "ai_response": ai_resp, "ocr_text": ocr_reusable_text if ocr_status == "success" else "",
                    "debug_info": {}
                })
                return

            # 阶段 2：LLM 摘要（流式输出思考过程）
            rag_context = ""
            rag_debug_details = _build_rag_debug_details(
                query="",
                job_title=job_title,
                difficulty=difficulty,
                interview_type=interview_type,
                style=style,
                stage="opening",
                status="not_started",
            )
            if ocr_status == "success" and ocr_raw_text:
                yield _sse_event("stage", {"stage": "AI 正在分析简历..."})
                q = f"请用简洁的中文总结以下简历的核心信息（姓名、技能、项目经历、教育背景），不要遗漏关键细节：\n\n{ocr_raw_text[:6000]}"
                if hasattr(agent, 'llm_client') and agent.llm_client:
                    messages = [{"role": "user", "content": q}]
                    for chunk in agent.llm_client.generate_stream(messages):
                        resume_summary += chunk
                        yield _sse_event("thinking", {"chunk": chunk})
                else:
                    resume_summary, steps = agent.run(
                        q,
                        trace_context=_build_langfuse_trace_context(
                            session_id,
                            interaction_type="opening.resume_summary",
                        ),
                    )
                    all_steps.extend(steps)

                if not resume_summary.strip():
                    fallback_summary = _build_resume_summary_fallback(ocr_raw_text)
                    if fallback_summary:
                        resume_summary = fallback_summary
                        all_steps.append({
                            "tool": "resume_summary_fallback",
                            "tool_input": "基于简历原文的确定性摘要回退",
                            "log": "流式摘要为空，回退到原文抽取摘要",
                            "observation": fallback_summary[:500],
                        })

                yield _sse_event("stage", {"stage": "正在检索知识库..."})
                rag_context, rag_debug_details = _retrieve_rag_context(
                    job_title=job_title,
                    difficulty=difficulty,
                    interview_type=interview_type,
                    style=style,
                    resume_text=ocr_raw_text,
                    stage="opening",
                )

            if explicit_resume_requested and has_resume and not resume_summary:
                yield _sse_event("error", {
                    "message": "简历文本已提取，但未能生成有效摘要，请重新上传内容更完整的简历。"
                })
                return

            # 阶段 3：注入 prompt + 开场白
            if hasattr(agent, 'update_dynamic_prompt'):
                generated_prompt = ""
                try:
                    generated_prompt = _try_generate_prompt(
                        job_title, interview_type, difficulty, style, resume_summary, job_requirements
                    )
                    if generated_prompt:
                        prompt_source = "prompt_generator"
                except Exception:
                    pass
                agent.update_dynamic_prompt(
                    job_title=job_title, interview_type=interview_type,
                    difficulty=difficulty, style=style,
                    feature_vad=feature_vad, feature_deep=feature_deep,
                    resume_summary=resume_summary, job_requirements=job_requirements,
                    custom_prompt=generated_prompt,
                    rag_context=rag_context
                )
            else:
                agent.update_system_prompt(style=style)
            agent.reset_memory()

            yield _sse_event("stage", {"stage": "AI 面试官正在准备开场白..."})
            if has_resume:
                setup_query = (
                    f"我应聘的岗位是【{job_title}】。你刚才已经解析了我的真实简历，现在面试正式开始。"
                    "请以面试官的身份用一段话向我打招呼，简述对我简历的第一印象，然后明确要求候选人做 2-3 分钟的自我介绍。"
                    "直接输出候选人最终会听到的话术，不要附带任何括号中的语气说明、动作说明或舞台指令。"
                    "切记：绝对不要自己编造任何经历！"
                )
                parse_result_text = "[OK] Resume parsed successfully, interviewer has locked real project details. Please introduce yourself first."
            else:
                setup_query = (
                    f"我应聘的岗位是【{job_title}】。我没有提供简历。"
                    "请以面试官的身份向我打招呼，并要求候选人先做自我介绍。"
                    "直接输出候选人最终会听到的话术，不要附带任何括号中的语气说明、动作说明或舞台指令。"
                )
                parse_result_text = "No resume provided"

            response, steps = agent.run(
                setup_query,
                trace_context=_build_langfuse_trace_context(
                    session_id,
                    interaction_type="opening.first_question",
                ),
            )
            response = _sanitize_assistant_response(response, agent)
            all_steps.extend(steps)
            if _has_self_intro_request(response):
                print(f"[OK] Verification passed (stream): AI requires candidate self-introduction")
            else:
                print(f"[FAIL] Verification failed (stream): AI did not require self-introduction")
                print(f"[AI] AI response: {response[:200]}")

            assistant_message = _append_visible_message(session_id, "assistant", response)
            _create_pending_turn(
                session_id,
                question_message=assistant_message,
                question_text=response,
                source="interviewer_llm",
                question_type="opening",
                difficulty=difficulty,
                rag_candidates=rag_debug_details.get("questions") if isinstance(rag_debug_details, dict) else [],
            )

            debug_info = _build_debug_info(agent, session_id)
            debug_info["intermediate_steps"] = all_steps
            debug_info["prompt_source"] = prompt_source
            debug_info["resume_summary"] = resume_summary[:2000]
            debug_info["ocr_raw_text"] = ocr_raw_text
            debug_info["ocr_status"] = ocr_status
            debug_info["rag_context"] = rag_context[:1000]
            debug_info["rag_details"] = rag_debug_details

            yield _sse_event("done", {
                "status": "success", "token": token,
                "session_id": session_id,
                "system_message": f"[OK] Interview room created (ID: {session_id[:8]})",
                "parse_result": parse_result_text,
                "ai_response": response,
                "ocr_text": ocr_reusable_text if ocr_status == "success" else "",
                "debug_info": debug_info
            })
        except Exception as e:
            traceback.print_exc()
            yield _sse_event("error", {"message": str(e)})

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/end-stream', methods=['POST'])
@require_session
def end_interview_stream(session_id):
    """流式结束面试：通过 SSE 实时输出评估思考过程。"""
    save_history = _parse_save_history_payload(default=True)
    agent = _agents.get(session_id)
    observer = None
    session_owner_id = _resolve_session_owner_id(session_id)

    # 创建 SSE 消息队列
    sse_queue = queue.Queue()

    # 设置观察者的 SSE 推送回调
    if observer:
        def push_callback(event):
            """SSE 推送回调：将评估更新推送到前端"""
            try:
                sse_queue.put(event)
            except Exception as e:
                print(f"[app] SSE 推送回调异常：{e}")
        observer.set_push_callback(push_callback)
    observer = _observers.pop(session_id, None)

    # 设置观察者的 SSE 推送回调
    if observer:
        def push_callback(event):
            """SSE 推送回调：将评估更新推送到前端"""
            try:
                # 直接 yield 事件（在 generate 函数内部）
                pass
            except Exception as e:
                print(f"[app] SSE 推送回调异常：{e}")
        observer.set_push_callback(push_callback)
        _retry_failed_turn_evaluations_before_end(session_id, observer, save_history=save_history)

    # 取出观察者草稿，清理 observer
    observer = observer
    draft = observer.shutdown(
        wait=save_history,
        timeout=EVAL_OBSERVER_DRAIN_TIMEOUT_SECONDS if save_history else None,
    ) if observer else {}
    
    if False and observer:
        # [NEW] Set SSE push callback
        def push_callback(event):
            yield _sse_event("eval_draft", event)
        
        observer.set_push_callback(push_callback)
        draft = observer.get_draft()
        observer.shutdown(wait=False)

    def generate():
        eval_result = {}
        if not save_history:
            yield _sse_event("stage", {"stage": "正在结束本次面试..."})

            if STORAGE_AVAILABLE:
                deleted = False
                if session_owner_id is not None:
                    deleted = data_client.delete_session(session_id, session_owner_id)
                if not deleted:
                    _skip_pending_turns_before_end(session_id)
                    data_client.end_session(session_id)

            _clear_session_runtime(session_id)
            revoke_session(session_id)

            payload = {
                "status": "success",
                "session_id": session_id,
                "saved": False,
                "report_available": False,
                "message": "本次面试未保存，评估报告不会生成。",
            }
            if session_owner_id is not None:
                payload["quota"] = _build_history_quota(session_owner_id)

            yield _sse_event("done", payload)
            return

        if agent:
            try:
                yield _sse_event("stage", {"stage": "AI 面试官正在撰写评估报告..."})
                # 流式评估
                if hasattr(agent, 'llm_client') and agent.llm_client:
                    eval_prompt = _build_eval_prompt(agent, draft=draft, session_id=session_id)
                    if eval_prompt:
                        messages = [{"role": "user", "content": eval_prompt}]
                        raw = ""
                        for chunk in agent.llm_client.generate_stream(messages):
                            raw += chunk
                            yield _sse_event("thinking", {"chunk": chunk})
                        eval_result = _parse_eval_result(raw)
                    else:
                        eval_result = agent.evaluate_interview(draft=draft)
                else:
                    eval_result = agent.evaluate_interview()
            except Exception as e:
                print(f"[end-stream] AI 评估失败: {e}")
        if not eval_result:
            eval_result = _build_fallback_eval_result_from_structured_data(session_id, draft)
        eval_result = _normalize_final_eval_result(eval_result, session_id=session_id, draft=draft)

        if STORAGE_AVAILABLE:
            _skip_pending_turns_before_end(session_id)
            data_client.end_session(session_id)
            ai_evals = eval_result.get("evaluations", [])
            if ai_evals:
                for ev in ai_evals:
                    data_client.save_evaluation(session_id, ev.get("dimension", "未知"), ev.get("score", 5), ev.get("comment", ""))
            else:
                existing_stats = data_client.get_session_statistics(session_id)
                if not existing_stats.get("evaluations"):
                    for dim, score, comment in [("技术深度", 7, "待评估"), ("沟通表达", 7, "待评估"), ("逻辑思维", 7, "待评估"), ("项目经验", 7, "待评估")]:
                        data_client.save_evaluation(session_id, dim, score, comment)
            stats = data_client.get_session_statistics(session_id)
            # 持久化评估总结文本
            data_client.save_eval_summary(
                session_id,
                strengths=eval_result.get("strengths", ""),
                weaknesses=eval_result.get("weaknesses", ""),
                summary=eval_result.get("summary", ""),
            )
        else:
            stats = None

        _clear_session_runtime(session_id)
        revoke_session(session_id)

        yield _sse_event("done", {
            "status": "success",
            "session_id": session_id,
            "saved": True,
            "report_available": True,
            "stats": stats,
            "strengths": eval_result.get("strengths", ""),
            "weaknesses": eval_result.get("weaknesses", ""),
            "summary": eval_result.get("summary", ""),
            **_final_report_response_fields(eval_result),
            "quota": _build_history_quota(session_owner_id) if session_owner_id is not None else None,
        })

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


def _build_structured_report_context(session_id: str) -> str:
    if not (STORAGE_AVAILABLE and data_client):
        return ""

    try:
        turns = data_client.list_interview_turns(session_id) if hasattr(data_client, "list_interview_turns") else []
        metadata_rows = data_client.list_question_metadata(session_id) if hasattr(data_client, "list_question_metadata") else []
        evaluation_rows = data_client.list_turn_evaluations(session_id) if hasattr(data_client, "list_turn_evaluations") else []
    except Exception as exc:
        print(f"[app] build structured report context failed: {exc}")
        return ""

    if not turns:
        return ""

    metadata_by_turn = {item.get("turn_id"): item for item in metadata_rows if isinstance(item, dict)}
    evaluations_by_turn = {}
    for item in evaluation_rows:
        if not isinstance(item, dict):
            continue
        evaluations_by_turn.setdefault(item.get("turn_id"), []).append(item)

    blocks = []
    for turn in turns:
        if not isinstance(turn, dict) or turn.get("status") != "answered":
            continue
        turn_id = turn.get("turn_id")
        metadata = metadata_by_turn.get(turn_id) or {}
        dimensions = metadata.get("dimensions") or []
        dimension_lines = []
        for dimension in dimensions[:2]:
            if not isinstance(dimension, dict):
                continue
            dimension_lines.append(
                f"- {dimension.get('name', '综合表现')}："
                f"rubric={dimension.get('rubric', '')[:160]}；"
                f"pass={dimension.get('pass_criteria', '')[:120]}"
            )
        evaluation_lines = []
        for evaluation in evaluations_by_turn.get(turn_id, [])[:3]:
            evaluation_lines.append(
                f"- {evaluation.get('dimension', '')}："
                f"{evaluation.get('score', 0)}分/{evaluation.get('pass_level', '')}；"
                f"证据：{evaluation.get('evidence', '')[:180]}；"
                f"建议：{evaluation.get('suggestion', '')[:120]}"
            )

        blocks.append(
            "\n".join([
                f"### 第 {turn.get('turn_no')} 轮",
                f"问题：{(turn.get('question_text') or '')[:500]}",
                f"回答：{(turn.get('answer_text') or '')[:700]}",
                "考察维度：",
                "\n".join(dimension_lines) if dimension_lines else "- 综合表现：通用面试表现",
                "过程评分：",
                "\n".join(evaluation_lines) if evaluation_lines else "- 暂无过程评分",
            ])
        )

    if not blocks:
        return ""

    return "【结构化面试记录】\n" + "\n\n".join(blocks)


def _build_fallback_eval_result_from_structured_data(session_id: str, draft: dict | None = None) -> dict:
    if not (STORAGE_AVAILABLE and data_client and hasattr(data_client, "list_turn_evaluations")):
        return {}

    try:
        rows = data_client.list_turn_evaluations(session_id)
    except Exception:
        rows = []

    if not rows:
        return {}

    grouped = {}
    for item in rows:
        if not isinstance(item, dict):
            continue
        dimension = str(item.get("dimension") or "").strip()
        if not dimension:
            continue
        grouped.setdefault(dimension, []).append(item)

    evaluations = []
    for dimension, items in grouped.items():
        scores = [int(item.get("score") or 0) for item in items if int(item.get("score") or 0) > 0]
        avg_score = round(sum(scores) / len(scores), 1) if scores else 0
        evidence = next((item.get("evidence") for item in items if item.get("evidence")), "")
        evaluations.append({
            "dimension": dimension,
            "score": avg_score,
            "comment": evidence or "基于过程评分聚合生成",
        })

    if not evaluations:
        return {}

    strengths = "；".join(item.get("text", "") for item in (draft or {}).get("strengths", []) if item.get("text"))
    weaknesses = "；".join(item.get("text", "") for item in (draft or {}).get("weaknesses", []) if item.get("text"))
    if not strengths:
        strengths = "候选人在部分维度中提供了可追溯的回答证据。"
    if not weaknesses:
        weaknesses = "建议结合过程评分中的追问建议继续补充薄弱细节。"

    avg = round(sum(item["score"] for item in evaluations) / len(evaluations), 1)
    return {
        "evaluations": evaluations,
        "strengths": strengths,
        "weaknesses": weaknesses,
        "summary": f"本报告基于 {len(rows)} 条过程评分聚合生成，平均分 {avg}。",
        "overall_score": avg,
        "hire_recommendation": _hire_recommendation_from_score(avg),
        "dimension_scores": _aggregate_dimension_scores_from_rows(rows),
        "evidence": _report_evidence_from_rows(rows),
        "next_training_plan": _derive_next_training_plan(
            _aggregate_dimension_scores_from_rows(rows),
            [weaknesses] if weaknesses else [],
        ),
        "source": "structured_fallback",
    }


def _normalize_final_eval_result(eval_result: dict, *, session_id: str = "", draft: dict | None = None) -> dict:
    if not isinstance(eval_result, dict) or not eval_result:
        return {}

    result = dict(eval_result)
    structured_rows = _safe_list_turn_evaluations(session_id)

    if structured_rows:
        dimension_scores = _aggregate_dimension_scores_from_rows(structured_rows)
        evidence = _report_evidence_from_rows(structured_rows)
    else:
        dimension_scores = _normalize_dimension_scores(
            result.get("dimension_scores") or result.get("evaluations") or []
        )
        evidence = _normalize_report_evidence(result.get("evidence") or [])

    if not dimension_scores:
        dimension_scores = _normalize_dimension_scores(result.get("evaluations") or [])
    if not evidence:
        evidence = _normalize_report_evidence(result.get("evidence") or [])

    legacy_evaluations = _legacy_evaluations_from_dimension_scores(dimension_scores)
    if not legacy_evaluations:
        legacy_evaluations = _normalize_legacy_evaluations(result.get("evaluations") or [])

    overall_score = _coerce_report_score(result.get("overall_score"))
    if not overall_score:
        score_values = [
            _coerce_report_score(item.get("score"))
            for item in dimension_scores or legacy_evaluations
            if isinstance(item, dict)
        ]
        score_values = [score for score in score_values if score]
        overall_score = round(sum(score_values) / len(score_values), 1) if score_values else 0

    strengths_items = _report_text_items(result.get("strengths"))
    weaknesses_items = _report_text_items(result.get("weaknesses"))
    if draft:
        if not strengths_items:
            strengths_items = _report_text_items(draft.get("strengths"))
        if not weaknesses_items:
            weaknesses_items = _report_text_items(draft.get("weaknesses"))

    next_training_plan = _report_text_items(result.get("next_training_plan"))
    if not next_training_plan:
        next_training_plan = _derive_next_training_plan(dimension_scores, weaknesses_items)

    summary = _report_text(result.get("summary"))
    if not summary and overall_score:
        summary = f"本次面试综合评分 {overall_score}/10，建议结合维度评分和证据继续复盘。"

    report = {
        "overall_score": overall_score,
        "hire_recommendation": _normalize_hire_recommendation(
            result.get("hire_recommendation"),
            overall_score,
        ),
        "dimension_scores": dimension_scores,
        "strengths": strengths_items,
        "weaknesses": weaknesses_items,
        "evidence": evidence,
        "next_training_plan": next_training_plan,
        "summary": summary,
    }

    result.update({
        "overall_score": report["overall_score"],
        "hire_recommendation": report["hire_recommendation"],
        "dimension_scores": report["dimension_scores"],
        "evidence": report["evidence"],
        "next_training_plan": report["next_training_plan"],
        "report": report,
        # Legacy compatibility for current frontend and persistence layer.
        "evaluations": legacy_evaluations,
        "strengths": "；".join(strengths_items),
        "weaknesses": "；".join(weaknesses_items),
        "summary": summary,
    })
    return result


def _final_report_response_fields(eval_result: dict) -> dict:
    if not isinstance(eval_result, dict):
        return {}
    return {
        "overall_score": eval_result.get("overall_score", 0),
        "hire_recommendation": eval_result.get("hire_recommendation", ""),
        "dimension_scores": eval_result.get("dimension_scores", []),
        "evidence": eval_result.get("evidence", []),
        "next_training_plan": eval_result.get("next_training_plan", []),
        "report": eval_result.get("report", {}),
    }


def _safe_list_turn_evaluations(session_id: str) -> list:
    if not (session_id and STORAGE_AVAILABLE and data_client and hasattr(data_client, "list_turn_evaluations")):
        return []
    try:
        rows = data_client.list_turn_evaluations(session_id) or []
    except Exception:
        return []
    return [row for row in rows if isinstance(row, dict)]


def _aggregate_dimension_scores_from_rows(rows: list) -> list:
    grouped = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        dimension = str(row.get("dimension") or "").strip()
        if not dimension:
            continue
        grouped.setdefault(dimension, []).append(row)

    dimension_scores = []
    for dimension, items in grouped.items():
        scores = [_coerce_report_score(item.get("score")) for item in items]
        scores = [score for score in scores if score]
        avg_score = round(sum(scores) / len(scores), 1) if scores else 0
        evidence = next((str(item.get("evidence") or "").strip() for item in items if item.get("evidence")), "")
        turns = _coerce_turn_list([item.get("turn_no") for item in items])
        dimension_scores.append({
            "dimension": dimension,
            "score": avg_score,
            "pass_level": _pass_level_from_report_score(avg_score),
            "evidence": evidence,
            "turns": turns,
            "comment": evidence or "基于过程评分聚合生成",
        })
    dimension_scores.sort(key=lambda item: item.get("score", 0), reverse=True)
    return dimension_scores


def _report_evidence_from_rows(rows: list) -> list:
    evidence_items = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        evidence = str(row.get("evidence") or "").strip()
        if not evidence:
            continue
        evidence_items.append({
            "turn_no": int(row.get("turn_no") or 0),
            "turn_id": str(row.get("turn_id") or ""),
            "dimension": str(row.get("dimension") or "综合表现"),
            "score": _coerce_report_score(row.get("score")),
            "pass_level": str(row.get("pass_level") or ""),
            "evidence": evidence,
            "suggestion": str(row.get("suggestion") or ""),
        })
    evidence_items.sort(key=lambda item: (item.get("turn_no") or 0, item.get("dimension") or ""))
    return evidence_items[:12]


def _normalize_dimension_scores(items) -> list:
    normalized = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        dimension = str(item.get("dimension") or "").strip()
        if not dimension:
            continue
        score = _coerce_report_score(item.get("score"))
        evidence = _report_text(item.get("evidence") or item.get("comment"))
        normalized.append({
            "dimension": dimension,
            "score": score,
            "pass_level": str(item.get("pass_level") or _pass_level_from_report_score(score)),
            "evidence": evidence,
            "turns": _coerce_turn_list(item.get("turns") or item.get("turn_no")),
            "comment": _report_text(item.get("comment") or evidence),
        })
    return normalized


def _normalize_report_evidence(items) -> list:
    normalized = []
    for item in items or []:
        if isinstance(item, str):
            text = item.strip()
            if text:
                normalized.append({"turn_no": 0, "dimension": "综合表现", "evidence": text})
            continue
        if not isinstance(item, dict):
            continue
        evidence = _report_text(item.get("evidence") or item.get("text") or item.get("comment"))
        if not evidence:
            continue
        normalized.append({
            "turn_no": int(item.get("turn_no") or item.get("turn") or 0),
            "turn_id": str(item.get("turn_id") or ""),
            "dimension": str(item.get("dimension") or "综合表现"),
            "score": _coerce_report_score(item.get("score")),
            "pass_level": str(item.get("pass_level") or ""),
            "evidence": evidence,
            "suggestion": str(item.get("suggestion") or ""),
        })
    return normalized[:12]


def _legacy_evaluations_from_dimension_scores(dimension_scores: list) -> list:
    evaluations = []
    for item in dimension_scores or []:
        if not isinstance(item, dict):
            continue
        dimension = str(item.get("dimension") or "").strip()
        if not dimension:
            continue
        evaluations.append({
            "dimension": dimension,
            "score": _coerce_report_score(item.get("score")) or 0,
            "comment": _report_text(item.get("comment") or item.get("evidence") or "基于结构化证据生成"),
        })
    return evaluations


def _normalize_legacy_evaluations(items) -> list:
    evaluations = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        dimension = str(item.get("dimension") or "").strip()
        if not dimension:
            continue
        evaluations.append({
            "dimension": dimension,
            "score": _coerce_report_score(item.get("score")) or 0,
            "comment": _report_text(item.get("comment") or item.get("evidence")),
        })
    return evaluations


def _derive_next_training_plan(dimension_scores: list, weaknesses: list) -> list:
    plan = []
    for item in dimension_scores or []:
        score = _coerce_report_score(item.get("score"))
        if score and score < 7:
            dimension = item.get("dimension") or "综合表现"
            plan.append(f"补强{dimension}：复盘相关项目，准备背景、动作、结果和取舍四类证据。")
        if len(plan) >= 3:
            break
    if not plan:
        for text in weaknesses or []:
            if text:
                plan.append(f"针对改进项复盘：{str(text)[:80]}")
            if len(plan) >= 3:
                break
    if not plan:
        plan.append("继续沉淀项目复盘材料，按问题背景、个人贡献、量化结果和技术取舍组织表达。")
    return plan


def _report_text_items(value) -> list:
    if not value:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, dict):
        text = _report_text(value.get("text") or value.get("summary") or value.get("comment"))
        return [text] if text else []
    items = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                text = _report_text(item.get("text") or item.get("summary") or item.get("comment"))
            else:
                text = _report_text(item)
            if text:
                items.append(text)
    return items


def _report_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        try:
            return json_mod.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value).strip()


def _coerce_report_score(value) -> float:
    try:
        score = float(value)
    except Exception:
        return 0
    if score <= 0:
        return 0
    return round(max(1, min(score, 10)), 1)


def _pass_level_from_report_score(score: float) -> str:
    if score >= 9:
        return "excellent"
    if score >= 7:
        return "pass"
    if score >= 5:
        return "weak_pass"
    if score > 0:
        return "fail"
    return ""


def _normalize_hire_recommendation(value, overall_score: float) -> str:
    text = str(value or "").strip()
    if text in {"recommend", "weak_recommend", "no_recommend"}:
        return text
    return _hire_recommendation_from_score(overall_score)


def _hire_recommendation_from_score(score: float) -> str:
    if score >= 8:
        return "recommend"
    if score >= 6:
        return "weak_recommend"
    return "no_recommend"


def _coerce_turn_list(value) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        values = value
    else:
        values = [value]
    turns = []
    for item in values:
        try:
            turn_no = int(item)
        except Exception:
            continue
        if turn_no > 0:
            turns.append(turn_no)
    return sorted(set(turns))


def _build_eval_prompt(agent, draft: dict = None, session_id: str = "") -> str:
    """构建评估 prompt，可注入观察者草稿作为参考"""
    try:
        history = agent.get_chat_history() if hasattr(agent, 'get_chat_history') else []
        structured_context = _build_structured_report_context(session_id) if session_id else ""
        if not history and not structured_context:
            return ""
        conversation = structured_context
        if not conversation:
            conversation = "\n".join([f"{'面试官' if m['role'] == 'assistant' else '候选人'}: {m['content']}" for m in history])

        # 构建草稿参考段落
        draft_context = ""
        if draft and (draft.get("strengths") or draft.get("weaknesses")):
            s_lines = "\n".join(
                f"  - 第{x['turn']}轮：{x['text']}" for x in draft.get("strengths", [])
            )
            w_lines = "\n".join(
                f"  - 第{x['turn']}轮：{x['text']}" for x in draft.get("weaknesses", [])
            )
            notes = "\n".join(
                f"  - 第{x['turn']}轮：{x['note']}" for x in draft.get("turn_notes", [])
            )
            draft_context = (
                f"\n\n【过程观察记录（供参考，请结合对话综合判断）】\n"
                f"优势观察：\n{s_lines}\n"
                f"不足观察：\n{w_lines}\n"
                f"关键节点：\n{notes}"
            )

        return f"""你是一位资深面试评估专家。请基于以下面试记录，生成一份专业的面试评估报告。
如果记录中包含过程评分，请优先复用其证据和维度，不要重复发明无依据的结论。
请输出证据型报告：每个维度都尽量关联具体轮次、证据和后续训练建议。

面试记录：
{conversation}{draft_context}

请严格按以下 JSON 格式输出（不要输出任何其他内容）：
{{
  "overall_score": 1-10,
  "hire_recommendation": "recommend | weak_recommend | no_recommend",
  "dimension_scores": [
    {{
      "dimension": "技术深度",
      "score": 1-10,
      "pass_level": "excellent | pass | weak_pass | fail",
      "evidence": "引用具体轮次的证据",
      "turns": [1, 2],
      "comment": "一句话评价"
    }}
  ],
  "evidence": [
    {{
      "turn_no": 1,
      "dimension": "技术深度",
      "score": 1-10,
      "pass_level": "excellent | pass | weak_pass | fail",
      "evidence": "候选人的具体回答证据",
      "suggestion": "后续改进或追问方向"
    }}
  ],
  "next_training_plan": [
    "针对薄弱项的训练建议"
  ],
  "evaluations": [
    {{"dimension": "技术深度", "score": 1-10, "comment": "评价"}},
    {{"dimension": "沟通表达", "score": 1-10, "comment": "评价"}},
    {{"dimension": "逻辑思维", "score": 1-10, "comment": "评价"}},
    {{"dimension": "项目经验", "score": 1-10, "comment": "评价"}},
    {{"dimension": "学习能力", "score": 1-10, "comment": "评价"}}
  ],
  "strengths": "候选人的主要优势（2-3句话）",
  "weaknesses": "候选人的不足之处（2-3句话）",
  "summary": "总体评价（3-5句话）"
}}"""
    except Exception:
        return ""


def _parse_eval_result(raw: str) -> dict:
    """解析评估 LLM 输出为结构化结果"""
    import re
    try:
        json_match = re.search(r'\{[\s\S]*\}', raw)
        if json_match:
            return json_mod.loads(json_match.group())
    except Exception:
        pass
    return {}


@app.route('/api/chat', methods=['POST'])
@require_session
def chat(session_id):
    """处理对话"""
    data = request.json
    user_message = data.get('message', '')
    if not user_message:
        return jsonify({"error": "消息不能为空"}), 400

    user_message_record = _append_visible_message(session_id, "user", user_message)
    answered_turn = _answer_pending_turn(
        session_id,
        answer_message=user_message_record,
        answer_text=user_message,
    )

    _chat_stop_flags[session_id] = False
    agent = _agents.get(session_id)
    debug_info = {}
    if agent:
        try:
            followup_context = _build_interviewer_hidden_context(session_id)
            response, steps = agent.run(
                user_message,
                context=followup_context,
                trace_context=_build_langfuse_trace_context(
                    session_id,
                    interaction_type="chat",
                    extra_metadata=_turn_trace_metadata(answered_turn),
                ),
            )
            response = _sanitize_assistant_response(response, agent)
            assistant_message = _append_visible_message(session_id, "assistant", response)
            _create_pending_turn(
                session_id,
                question_message=assistant_message,
                question_text=response,
                source="interviewer_llm",
                question_type="followup",
            )
            debug_info = _build_debug_info(agent, session_id)
            debug_info["intermediate_steps"] = steps
            # 后台异步触发观察者分析本轮
            observer = _observers.get(session_id)
            if observer:
                accepted = observer.observe_async(answered_turn) if answered_turn else False
                if not accepted:
                    observer.observe_async(agent.get_chat_history())
        except Exception as e:
            _record_agent_event(
                session_id,
                "assistant_reply_failed",
                turn_id=(answered_turn or {}).get("turn_id", ""),
                payload={"error": str(e)[:200]},
            )
            response = f"抱歉，系统遇到了问题：{str(e)}"
    else:
        import time
        time.sleep(1.5)
        response = f"针对你刚才说的\u201c{user_message[:10]}...\u201d，你能详细谈谈底层的实现原理吗？"
        response = _sanitize_assistant_response(response)
        assistant_message = _append_visible_message(session_id, "assistant", response)
        _create_pending_turn(
            session_id,
            question_message=assistant_message,
            question_text=response,
            source="mock",
            question_type="followup",
        )
        observer = _observers.get(session_id)
        if observer and answered_turn:
            observer.observe_async(answered_turn)

    return jsonify({"response": response, "debug_info": debug_info})


@app.route('/api/chat-stop', methods=['POST'])
@require_session
def stop_chat_stream(session_id):
    """Best-effort stop signal for the current streaming reply."""
    _chat_stop_flags[session_id] = True
    return jsonify({"status": "success"})


@app.route('/api/chat-stream', methods=['POST'])
@require_session
def chat_stream(session_id):
    """流式对话：通过 SSE 实时输出 LLM 思维链"""
    data = request.json
    user_message = data.get('message', '')
    if not user_message:
        return jsonify({"error": "消息不能为空"}), 400

    user_message_record = _append_visible_message(session_id, "user", user_message)
    answered_turn = _answer_pending_turn(
        session_id,
        answer_message=user_message_record,
        answer_text=user_message,
    )

    agent = _agents.get(session_id)
    observer = _observers.get(session_id)
    followup_context = _build_interviewer_hidden_context(session_id)

    # 创建 SSE 消息队列
    sse_queue = queue.Queue()

    # 设置观察者的 SSE 推送回调
    if observer:
        def push_callback(event):
            """SSE 推送回调：将评估更新推送到前端"""
            try:
                sse_queue.put(event)
            except Exception as e:
                print(f"[app] SSE 推送回调异常：{e}")
        observer.set_push_callback(push_callback)

    def generate():
        full_response = ""
        interrupted = False
        stream_sanitizer = _StreamStageDirectionSanitizer()
        if agent and hasattr(agent, 'run_stream'):
            yield _sse_event("stage", {"stage": "面试官正在深度思考..."})
            try:
                in_thinking = False
                for chunk_type, chunk in agent.run_stream(user_message, context=followup_context):
                    if _chat_stop_flags.get(session_id):
                        interrupted = True
                        break
                    if chunk_type == "thinking":
                        if not in_thinking:
                            in_thinking = True
                        yield _sse_event("thinking", {"chunk": chunk})
                    elif chunk_type == "content":
                        if in_thinking:
                            # 思维链结束，正式回复开始
                            in_thinking = False
                            yield _sse_event("stage", {"stage": "面试官正在组织回复..."})
                        cleaned_chunk = stream_sanitizer.feed(chunk)
                        if cleaned_chunk:
                            full_response += cleaned_chunk
                            yield _sse_event("content", {"chunk": cleaned_chunk})
                    
                    # 检查是否有评估更新需要推送
                    try:
                        while True:
                            eval_event = sse_queue.get_nowait()
                            yield _sse_event(eval_event["type"], eval_event["data"])
                    except queue.Empty:
                        pass
            except Exception as e:
                _record_agent_event(
                    session_id,
                    "assistant_reply_failed",
                    turn_id=(answered_turn or {}).get("turn_id", ""),
                    payload={"error": str(e)[:200], "mode": "chat_stream"},
                )
                full_response = f"抱歉，系统遇到了问题：{str(e)}"
        elif agent:
            # 不支持流式，降级为一次性调用
            yield _sse_event("stage", {"stage": "面试官正在思考..."})
            try:
                full_response, _ = agent.run(
                    user_message,
                    context=followup_context,
                    trace_context=_build_langfuse_trace_context(
                        session_id,
                        interaction_type="chat_stream.fallback",
                        extra_metadata=_turn_trace_metadata(answered_turn),
                    ),
                )
            except Exception as e:
                _record_agent_event(
                    session_id,
                    "assistant_reply_failed",
                    turn_id=(answered_turn or {}).get("turn_id", ""),
                    payload={"error": str(e)[:200], "mode": "chat_stream_fallback"},
                )
                full_response = f"抱歉，系统遇到了问题：{str(e)}"
        else:
            import time
            time.sleep(1.5)
            full_response = f"针对你刚才说的\u201c{user_message[:10]}...\u201d，你能详细谈谈底层的实现原理吗？"

        trailing_chunk = stream_sanitizer.flush()
        if trailing_chunk:
            full_response += trailing_chunk
            yield _sse_event("content", {"chunk": trailing_chunk})

        full_response = _sanitize_assistant_response(full_response, agent)

        assistant_message = None
        if full_response:
            assistant_message = _append_visible_message(session_id, "assistant", full_response)
            if interrupted:
                _record_agent_event(
                    session_id,
                    "assistant_reply_interrupted",
                    turn_id=(answered_turn or {}).get("turn_id", ""),
                    payload={"response_preview": full_response[:160]},
                )
            elif not _is_assistant_error_response(full_response):
                _create_pending_turn(
                    session_id,
                    question_message=assistant_message,
                    question_text=full_response,
                    source="interviewer_llm" if agent else "mock",
                    question_type="followup",
                )
        if interrupted and agent and full_response and hasattr(agent, '_add_to_history'):
            try:
                agent._add_to_history({"role": "user", "content": user_message})
                agent._add_to_history({"role": "assistant", "content": full_response})
            except Exception:
                pass

        # 后台异步触发观察者分析本轮
        if agent and full_response:
            observer = _observers.get(session_id)
            if observer:
                accepted = observer.observe_async(answered_turn) if answered_turn else False
                if not accepted:
                    observer.observe_async(agent.get_chat_history())

        # 推送所有剩余的评估更新
        while True:
            try:
                eval_event = sse_queue.get_nowait()
                yield _sse_event(eval_event["type"], eval_event["data"])
            except queue.Empty:
                break

        debug_info = _build_debug_info(agent, session_id) if agent else {}
        yield _sse_event("done", {
            "response": full_response,
            "interrupted": interrupted,
            "debug_info": debug_info,
        })

        _chat_stop_flags.pop(session_id, None)

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/api/end', methods=['POST'])
@require_session
def end_interview(session_id):
    """结束面试"""
    save_history = _parse_save_history_payload(default=True)
    agent = _agents.get(session_id)
    eval_result = {}
    session_owner_id = _resolve_session_owner_id(session_id)

    # 取出观察者草稿，清理 observer
    observer = _observers.pop(session_id, None)
    _retry_failed_turn_evaluations_before_end(session_id, observer, save_history=save_history)
    draft = (
        observer.shutdown(
            wait=save_history,
            timeout=EVAL_OBSERVER_DRAIN_TIMEOUT_SECONDS if save_history else None,
        )
        if observer else {}
    )

    if not save_history:
        if STORAGE_AVAILABLE:
            deleted = False
            if session_owner_id is not None:
                deleted = data_client.delete_session(session_id, session_owner_id)
            if not deleted:
                _skip_pending_turns_before_end(session_id)
                data_client.end_session(session_id)

        _clear_session_runtime(session_id)
        revoke_session(session_id)

        payload = {
            "status": "success",
            "session_id": session_id,
            "saved": False,
            "report_available": False,
            "message": "本次面试未保存，评估报告不会生成。",
        }
        if session_owner_id is not None:
            payload["quota"] = _build_history_quota(session_owner_id)
        return jsonify(payload)

    # 调用 AI 生成真实评估（优先注入结构化 turn/metadata/evaluation，再回退到 agent 内存）
    if agent:
        try:
            eval_prompt = _build_eval_prompt(agent, draft=draft, session_id=session_id)
            if eval_prompt and hasattr(agent, 'llm_client') and agent.llm_client:
                eval_result = _parse_eval_result(agent.llm_client.generate([{"role": "user", "content": eval_prompt}]))
            else:
                eval_result = agent.evaluate_interview(draft=draft)
            print(f"[end] AI 评估完成: {list(eval_result.keys())}")
        except Exception as e:
            print(f"[end] AI 评估失败: {e}")
    if not eval_result:
        eval_result = _build_fallback_eval_result_from_structured_data(session_id, draft)
    eval_result = _normalize_final_eval_result(eval_result, session_id=session_id, draft=draft)

    if STORAGE_AVAILABLE:
        _skip_pending_turns_before_end(session_id)
        data_client.end_session(session_id)

        # 保存 AI 评估结果（如果有）
        ai_evals = eval_result.get("evaluations", [])
        if ai_evals:
            for ev in ai_evals:
                data_client.save_evaluation(
                    session_id,
                    ev.get("dimension", "未知"),
                    ev.get("score", 5),
                    ev.get("comment", "")
                )
        else:
            # fallback：占位评分
            existing_stats = data_client.get_session_statistics(session_id)
            if not existing_stats.get("evaluations"):
                placeholder_evals = [
                    ("技术深度", 7, "待 AI 评估模块接入后自动生成"),
                    ("沟通表达", 7, "待 AI 评估模块接入后自动生成"),
                    ("逻辑思维", 7, "待 AI 评估模块接入后自动生成"),
                    ("项目经验", 7, "待 AI 评估模块接入后自动生成"),
                ]
                for dim, score, comment in placeholder_evals:
                    data_client.save_evaluation(session_id, dim, score, comment)

        stats = data_client.get_session_statistics(session_id)
        # 持久化评估总结文本
        data_client.save_eval_summary(
            session_id,
            strengths=eval_result.get("strengths", ""),
            weaknesses=eval_result.get("weaknesses", ""),
            summary=eval_result.get("summary", ""),
        )
        revoke_session(session_id)
        _clear_session_runtime(session_id)
        return jsonify({
            "status": "success",
            "session_id": session_id,
            "saved": True,
            "report_available": True,
            "stats": stats,
            "strengths": eval_result.get("strengths", ""),
            "weaknesses": eval_result.get("weaknesses", ""),
            "summary": eval_result.get("summary", ""),
            **_final_report_response_fields(eval_result),
            "quota": _build_history_quota(session_owner_id) if session_owner_id is not None else None,
        })

    _clear_session_runtime(session_id)
    revoke_session(session_id)
    return jsonify({
        "status": "success",
        "session_id": session_id,
        "saved": True,
        "report_available": bool(eval_result),
        "message": "会话已结束",
        "strengths": eval_result.get("strengths", ""),
        "weaknesses": eval_result.get("weaknesses", ""),
        "summary": eval_result.get("summary", ""),
        **_final_report_response_fields(eval_result),
    })


def _sse_event(event: str, data: dict) -> str:
    """Format one SSE event payload."""
    return f"event: {event}\ndata: {json_mod.dumps(data, ensure_ascii=False)}\n\n"


def _parse_report_context(raw_value):
    if not raw_value:
        return None
    if isinstance(raw_value, dict):
        return raw_value
    if isinstance(raw_value, str):
        try:
            parsed = json_mod.loads(raw_value)
            return parsed if isinstance(parsed, dict) else None
        except Exception:
            return None
    return None


def _merge_job_title_with_report_context(job_title: str, report_context: dict | None) -> str:
    if not report_context:
        return job_title

    parts = []
    if job_title:
        parts.append(job_title)
    if report_context.get("position"):
        parts.append(report_context["position"])

    summary_bits = []
    if report_context.get("summary"):
        summary_bits.append(f"Interview summary: {report_context['summary']}")
    if report_context.get("strengths"):
        summary_bits.append(f"Strengths: {report_context['strengths']}")
    if report_context.get("weaknesses"):
        summary_bits.append(f"Weaknesses: {report_context['weaknesses']}")

    evaluations = report_context.get("evaluations") or []
    if evaluations:
        eval_text = "; ".join(
            f"{item.get('dimension', 'unknown')} {item.get('score', '')}/10 {item.get('comment', '')}".strip()
            for item in evaluations
        )
        if eval_text:
            summary_bits.append(f"Interview dimensions: {eval_text}")

    merged_job_title = " ".join(part for part in parts if part).strip()
    if not summary_bits:
        return merged_job_title
    if not merged_job_title:
        return "\n".join(summary_bits)
    return f"{merged_job_title}\n" + "\n".join(summary_bits)


@app.route('/api/resume/analyze-stream', methods=['POST'])
def analyze_resume_stream():
    """流式简历分析：通过 SSE 实时输出 LLM 思考过程。
    支持两种模式：
    1. FormData 上传文件 → 自动提取文本（OCR / 直读）+ 流式分析
    2. JSON body { ocr_text, job_title } → 跳过文件解析，直接流式分析
    """
    if not ANALYZER_AVAILABLE:
        return jsonify({"status": "error", "message": "简历分析模块不可用"}), 503

    # 解析请求参数
    if request.is_json:
        data = request.json or {}
        ocr_text = data.get('ocr_text', '')
        job_title = data.get('job_title', '')
        report_context = _parse_report_context(data.get('report_context'))
        if not ocr_text:
            return jsonify({"status": "error", "message": "ocr_text 不能为空"}), 400
        ocr_images = {}
    else:
        if 'resume' not in request.files:
            return jsonify({"status": "error", "message": "请上传简历文件"}), 400
        file = request.files['resume']
        if file.filename == '':
            return jsonify({"status": "error", "message": "文件名为空"}), 400
        job_title = request.form.get('job_title', '')
        report_context = _parse_report_context(request.form.get('report_context'))
        try:
            _, file_path = _save_uploaded_resume(file)
            extraction = _extract_resume_payload(file_path, include_images=True)
            ocr_text = extraction["text"] if extraction["success"] else extraction["error_message"]
            ocr_images = extraction["images"]
            if not extraction["success"]:
                return jsonify({"status": "error", "message": f"简历解析失败: {extraction['error_message']}"}), 400
        except ResumeOcrUnavailableError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 503
        except ResumeExtractionError as exc:
            return jsonify({"status": "error", "message": str(exc)}), 400
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    # 获取模型配置
    provider = get_default_provider()
    _is_json = request.is_json  # 在请求上下文内捕获，generator 内不能访问 request

    def generate():
        try:
            yield _sse_event("stage", {"stage": "简历解析完成，开始 AI 分析..."})
            # 把 OCR 原文摘要发给前端，让用户看到输入大模型的文字
            ocr_preview = ocr_text[:800] + ("..." if len(ocr_text) > 800 else "")
            yield _sse_event("thinking", {"chunk": f"[DOC] Resume original text (input to LLM):\n{ocr_preview}\n\n---\n\n"})

            analyzer = ResumeAnalyzer(
                api_key=provider.api_key,
                base_url=provider.base_url,
                model=provider.model,
            )
            analysis_job_title = _merge_job_title_with_report_context(job_title, report_context)
            final_result = None
            for event in analyzer.analyze_stream(ocr_text=ocr_text, job_title=analysis_job_title, report_context=report_context):
                if event["type"] == "stage":
                    yield _sse_event("stage", {"stage": event["stage"]})
                elif event["type"] == "thinking":
                    yield _sse_event("thinking", {"chunk": event["chunk"]})
                elif event["type"] == "done":
                    final_result = event["result"]

            # 创建 session token
            session_id = str(uuid.uuid4())
            token = create_token(session_id)
            _resume_sessions[token] = {
                "ocr_text": ocr_text,
                "sections": final_result.get("sections", []) if final_result else [],
                "suggestions": final_result.get("suggestions", []) if final_result else [],
            }

            yield _sse_event("done", {
                "status": "success",
                "token": token,
                "ocr_text": ocr_text[:3000],
                "sections": final_result.get("sections", []) if final_result else [],
                "suggestions": final_result.get("suggestions", []) if final_result else [],
                "builder_data": final_result.get("builder_data") if final_result else None,
                "images": ocr_images if not _is_json else {},
            })
        except Exception as e:
            traceback.print_exc()
            yield _sse_event("error", {"message": str(e)})

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


# ==========================================
# 简历优化模块 API
# ==========================================

# 简历分析会话数据（token -> 分析结果）
_resume_sessions: dict[str, dict] = {}


@app.route('/api/resume/analyze', methods=['POST'])
def analyze_resume():
    """上传简历，提取文本并执行 DeepSeek 分析，返回结构化建议。
    支持两种模式：
    1. FormData 上传文件 → 自动提取文本（OCR / 直读）+ 分析
    2. JSON body { ocr_text, job_title } → 跳过文件解析，直接分析（复用面试 OCR 结果）
    """
    if not ANALYZER_AVAILABLE:
        return jsonify({"status": "error", "message": "简历分析模块不可用"}), 503

    # 模式 2：JSON body 带 ocr_text，跳过文件上传和 OCR
    if request.is_json:
        data = request.json or {}
        ocr_text = data.get('ocr_text', '')
        job_title = data.get('job_title', '')
        report_context = _parse_report_context(data.get('report_context'))
        if not ocr_text:
            return jsonify({"status": "error", "message": "ocr_text 不能为空"}), 400

        try:
            analyzer = ResumeAnalyzer(
                api_key=DEEPSEEK_API_KEY,
                base_url=DEEPSEEK_BASE_URL
            )
            analysis_job_title = _merge_job_title_with_report_context(job_title, report_context)
            result = analyzer.analyze(ocr_text=ocr_text, job_title=analysis_job_title, report_context=report_context)

            session_id = str(uuid.uuid4())
            token = create_token(session_id)
            _resume_sessions[token] = {
                "ocr_text": ocr_text,
                "sections": result.get("sections", []),
                "suggestions": result.get("suggestions", [])
            }

            return jsonify({
                "status": "success",
                "token": token,
                "ocr_text": ocr_text[:3000],
                "sections": result.get("sections", []),
                "suggestions": result.get("suggestions", []),
                "builder_data": result.get("builder_data"),
                "images": {}
            })
        except Exception as e:
            traceback.print_exc()
            return jsonify({"status": "error", "message": str(e)}), 500

    # 模式 1：FormData 上传文件 → 自动提取文本 + 分析
    # 接收文件
    if 'resume' not in request.files:
        return jsonify({"status": "error", "message": "请上传简历文件"}), 400
    file = request.files['resume']
    if file.filename == '':
        return jsonify({"status": "error", "message": "文件名为空"}), 400

    job_title = request.form.get('job_title', '')
    report_context = _parse_report_context(request.form.get('report_context'))
    try:
        # 1. 提取简历文本（OCR / 直读自动选择）
        _, file_path = _save_uploaded_resume(file)
        extraction = _extract_resume_payload(file_path, include_images=True)
        ocr_result = extraction["text"] if extraction["success"] else extraction["error_message"]
        ocr_images = extraction["images"]

        if not extraction["success"]:
            return jsonify({"status": "error", "message": f"简历解析失败: {extraction['error_message']}"}), 400

        # 2. DeepSeek 分析
        analyzer = ResumeAnalyzer(
            api_key=DEEPSEEK_API_KEY,
            base_url=DEEPSEEK_BASE_URL
        )
        analysis_job_title = _merge_job_title_with_report_context(job_title, report_context)
        result = analyzer.analyze(ocr_text=ocr_result, job_title=analysis_job_title, report_context=report_context)

        # 3. 创建 session 并返回 token
        session_id = str(uuid.uuid4())
        token = create_token(session_id)
        _resume_sessions[token] = {
            "ocr_text": ocr_result,
            "sections": result.get("sections", []),
            "suggestions": result.get("suggestions", [])
        }

        return jsonify({
            "status": "success",
            "token": token,
            "ocr_text": ocr_result[:3000],
            "sections": result.get("sections", []),
            "suggestions": result.get("suggestions", []),
            "builder_data": result.get("builder_data"),
            "images": ocr_images
        })

    except ResumeOcrUnavailableError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 503
    except ResumeExtractionError as exc:
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/resume/export', methods=['POST'])
@require_session
def export_resume_pdf(session_id):
    """接收最终 sections JSON，用 Playwright 渲染 A4 PDF"""
    print(f"[INFO] Received export request, session_id: {session_id}")
    data = request.json
    sections = data.get('sections', [])
    print(f"[DOC] sections count: {len(sections)}")

    if not sections:
        return jsonify({"status": "error", "message": "sections 不能为空"}), 400

    try:
        print("[...] Importing dependencies...")
        from services.markdown_service import sections_to_html
        from services.pdf_service import render_resume_pdf
        print("[OK] Dependencies imported")

        # sections → structured HTML (properly handling markdown content)
        print("[...] Converting sections to HTML...")
        html_body = sections_to_html(sections)
        print(f"[OK] HTML generated, length: {len(html_body)}")

        # HTML → Playwright → PDF
        print("[...] Rendering PDF...")
        pdf_path = render_resume_pdf(html_body)
        print(f"[OK] PDF generated: {pdf_path}")

        return send_file(pdf_path, mimetype='application/pdf',
                         as_attachment=True, download_name='optimized_resume.pdf')

    except ImportError as ie:
        print(f"[FAIL] Dependency import failed: {ie}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"PDF export dependency not installed: {ie}"}), 503
    except Exception as e:
        print(f"[FAIL] Export failed: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/resume/polish', methods=['POST'])
def polish_resume_builder():
    """
    简历生成器 AI 优化端点
    接收简历文档的模块数据，返回优化建议
    """
    if not ANALYZER_AVAILABLE:
        return jsonify({"status": "error", "message": "简历分析模块不可用"}), 503

    data = request.json
    modules = data.get('modules', [])
    target_jd = data.get('targetJd', '')
    mode = data.get('mode', 'general')

    if not modules:
        return jsonify({"status": "error", "message": "modules 不能为空"}), 400

    try:
        # 初始化分析器
        analyzer = ResumeAnalyzer(
            api_key=DEEPSEEK_API_KEY,
            base_url=DEEPSEEK_BASE_URL,
            model="deepseek-chat"
        )

        # 将模块转换为文本格式供 LLM 分析
        resume_text = _convert_modules_to_text(modules)

        # 构建优化 prompt
        job_context = f"目标岗位：{target_jd}\n\n" if mode == 'targeted' and target_jd else ""

        prompt = f"""{job_context}你是一位资深的简历优化专家。请审查以下简历内容，找出可以优化的地方并给出具体的改写建议。

## 优化原则

1. **量化成果**：添加具体数字、百分比、规模等量化指标
2. **强化动词**：将"负责"、"参与"等弱动词改为"主导"、"优化"、"实现"等强动词
3. **STAR 法则**：确保每条经历包含情境(Situation)、任务(Task)、行动(Action)、结果(Result)
4. **技术关键词**：补充相关技术栈、工具、方法论的关键词
5. **简洁明确**：去除冗余表述，突出核心亮点

## 输出格式

对每个发现的优化点，输出一个 JSON 对象：
- "id": 唯一标识，格式 "sug_001", "sug_002" 等
- "moduleId": 对应模块的 id
- "entryId": 如果是时间线条目，提供 entry 的 id；否则为 null
- "fieldPath": 字段路径，如 "detail"、"content" 等
- "originalText": 原始文本片段（不超过 200 字）
- "suggestedText": 优化后的文本
- "reason": 优化理由（简洁说明，不超过 50 字）
- "status": 固定为 "pending"

请严格输出 JSON 数组，不要输出任何其他内容。不要用 markdown 代码块包裹。
如果某个模块没有明显可优化的地方，就不要为它生成建议。
最多生成 10 条建议，优先选择影响最大的优化点。

简历内容：
{resume_text}"""

        messages = [
            {"role": "system", "content": "你是简历优化专家，只输出合法 JSON 数组。"},
            {"role": "user", "content": prompt}
        ]

        # 调用 LLM
        raw_response = analyzer.client.generate(messages)
        suggestions = analyzer._parse_json_array(raw_response, fallback_id_prefix="sug")

        return jsonify({
            "status": "success",
            "suggestions": suggestions
        })

    except Exception as e:
        error_details = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "traceback": traceback.format_exc()
        }
        print(f"[FAIL] Resume optimization failed: {error_details}")
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e),
            "details": error_details
        }), 500


def _convert_modules_to_text(modules: list) -> str:
    """将前端模块数据转换为可读文本格式"""
    lines = []

    for mod in modules:
        if not mod.get('visible', True):
            continue

        lines.append(f"\n## {mod.get('title', '未命名模块')} (ID: {mod.get('id', 'unknown')})")

        # 求职意向
        if mod.get('type') == 'intention' and mod.get('intention'):
            intention = mod['intention']
            parts = [
                intention.get('targetJob'),
                intention.get('targetCity'),
                intention.get('salary'),
                intention.get('availableDate')
            ]
            lines.append(' | '.join(filter(None, parts)))

        # 时间线条目（教育、工作、项目等）
        elif mod.get('entries'):
            for entry in mod['entries']:
                time_range = f"{entry.get('timeStart', '')} - {'至今' if entry.get('isCurrent') else entry.get('timeEnd', '')}"
                lines.append(f"\n### {entry.get('orgName', '')} | {entry.get('role', '')} | {time_range}")
                lines.append(f"(Entry ID: {entry.get('id', 'unknown')})")
                if entry.get('detail'):
                    lines.append(entry['detail'])

        # 技能条
        elif mod.get('skillBars'):
            skills = [f"{sk.get('name', '')} ({sk.get('level', 0)}%)" for sk in mod['skillBars']]
            lines.append(' · '.join(skills))

        # 标签
        elif mod.get('tags'):
            lines.append('、'.join(mod['tags']))

        # 纯文本内容
        elif mod.get('content'):
            lines.append(mod['content'])

    return '\n'.join(lines)


@app.route('/api/export-pdf', methods=['POST'])
def export_markdown_pdf():
    """通用 Markdown → PDF 导出端点"""
    data = request.json
    md_text = data.get('markdown', '')
    if not md_text:
        return jsonify({"status": "error", "message": "markdown 不能为空"}), 400

    try:
        from services.markdown_service import convert_markdown_to_html
        from services.pdf_service import render_resume_pdf

        html_body = convert_markdown_to_html(md_text)
        pdf_path = render_resume_pdf(html_body)
        filename = os.path.basename(pdf_path)

        return jsonify({"status": "success", "pdf_url": f"/download/{filename}"})

    except ImportError as ie:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"依赖未安装: {ie}"}), 503
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/export-html-pdf', methods=['POST'])
def export_html_pdf():
    """HTML → PDF 导出端点（用于简历生成器）"""
    data = request.json
    html_content = data.get('html', '')
    if not html_content:
        return jsonify({"status": "error", "message": "html 不能为空"}), 400

    try:
        from services.pdf_service import render_html_to_pdf

        # 直接渲染 HTML 为 PDF
        pdf_path = render_html_to_pdf(html_content)

        return send_file(pdf_path, mimetype='application/pdf',
                         as_attachment=True, download_name='resume.pdf')

    except ImportError as ie:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"依赖未安装: {ie}"}), 503
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/download/<filename>', methods=['GET'])
def download_pdf(filename):
    """下载已生成的 PDF 文件"""
    generated_dir = PDF_OUTPUT_DIR
    file_path = os.path.join(generated_dir, filename)
    if not os.path.exists(file_path):
        return jsonify({"status": "error", "message": "文件不存在"}), 404
    return send_file(file_path, mimetype='application/pdf', as_attachment=True)


# ==========================================
# 语音交互模块 API
# ==========================================

@app.route('/api/speech/stt', methods=['POST'])
@require_session
def speech_to_text(session_id):
    """语音识别：接收前端录制的 PCM/WAV 音频，返回识别文本"""
    if not SPEECH_AVAILABLE:
        return jsonify({"status": "error", "message": "语音服务未配置"}), 503

    if 'audio' not in request.files:
        return jsonify({"status": "error", "message": "缺少 audio 文件"}), 400

    audio_file = request.files['audio']
    audio_data = audio_file.read()
    if not audio_data:
        return jsonify({"status": "error", "message": "音频数据为空"}), 400

    # 前端 MediaRecorder 录制的是 webm，需要判断格式
    fmt = request.form.get('format', 'pcm')
    rate = int(request.form.get('rate', '16000'))

    try:
        text = _speech_client.speech_to_text(audio_data, fmt=fmt, rate=rate)
        return jsonify({"status": "success", "text": text})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/speech/polish', methods=['POST'])
@require_session
def polish_speech_text(session_id):
    """语音文本清洗：用 LLM 纠正语音识别中的技术术语错误。
    返回 { text: 纠正后文本, corrections: [{original, corrected, start, end}] }
    """
    data = request.json or {}
    raw_text = data.get('text', '')
    if not raw_text:
        return jsonify({"status": "error", "message": "文本为空"}), 400

    agent = _agents.get(session_id)
    if not agent or not agent.llm_client:
        return jsonify({"status": "success", "text": raw_text, "corrections": []})

    polish_prompt = """你是一个语音识别纠错助手。用户通过语音输入回答技术面试问题，但语音识别对英文技术术语的准确率很低，经常把英文词汇音译成中文。

请纠正以下文本中的技术术语错误，只修正明显的语音误识别，不要改变用户的原意和表达方式。

严格按 JSON 格式输出（不要输出任何其他内容）：
{"text": "纠正后的完整文本", "corrections": [{"original": "错误词", "corrected": "正确词"}]}

如果没有需要纠正的内容，corrections 返回空数组。

常见误识别参考：福莱克斯→Flask, 姜戈→Django, 瑞艾克特→React, 派森→Python, 加瓦→Java, 西加加→C++, 多克→Docker, 瑞迪斯→Redis, 卡夫卡→Kafka, 恩金艾克斯→Nginx, 库伯奈提斯→Kubernetes, 弗拉斯克→Flask, 泰森弗洛→TensorFlow, 派托奇→PyTorch, 麦艾斯克→MySQL, 蒙戈→MongoDB, 诺的→Node, 威优异→Vue, 斯普林→Spring, 盖特→Git, 艾皮艾→API, 杰森→JSON, 艾奇提提皮→HTTP"""

    messages = [
        {"role": "system", "content": polish_prompt},
        {"role": "user", "content": raw_text}
    ]

    try:
        import re
        result = agent.llm_client.generate(messages)
        json_match = re.search(r'\{[\s\S]*\}', result)
        if json_match:
            import json
            parsed = json.loads(json_match.group())
            return jsonify({
                "status": "success",
                "text": parsed.get("text", raw_text),
                "corrections": parsed.get("corrections", [])
            })
    except Exception as e:
        print(f"[polish] 语音纠错失败: {e}")

    return jsonify({"status": "success", "text": raw_text, "corrections": []})

@app.route('/api/speech/tts', methods=['POST'])
@require_session
def text_to_speech(session_id):
    """语音合成：接收文本，返回 PCM 音频二进制"""
    return _do_tts()


@app.route('/api/speech/tts-preview', methods=['POST'])
def tts_preview():
    """语音试听（无需认证，供配置页使用）"""
    return _do_tts()


@app.route('/api/debug/sysprompt', methods=['GET'])
def get_current_sysprompt():
    """获取当前 session 正在使用的 System Prompt（调试用途）"""
    session_id = request.args.get('session_id')
    
    if not session_id:
        return jsonify({
            "status": "error",
            "message": "缺少 session_id 参数"
        }), 400
    
    if session_id not in _agents:
        return jsonify({
            "status": "error",
            "message": "该 session 不存在或已过期"
        }), 404
    
    agent = _agents[session_id]
    
    # 获取 agent 的 current_prompt
    if hasattr(agent, 'prompt'):
        sysprompt = agent.prompt
        current_style = getattr(agent, 'current_style', 'unknown')
        
        return jsonify({
            "status": "success",
            "session_id": session_id,
            "sysprompt": sysprompt,
            "current_style": current_style,
            "prompt_length": len(sysprompt),
            "estimated_tokens": len(sysprompt) // 4
        })
    else:
        return jsonify({
            "status": "error",
            "message": "Agent 未初始化 prompt"
        }), 500


def _do_tts():
    """TTS 公共逻辑"""
    if not SPEECH_AVAILABLE:
        return jsonify({"status": "error", "message": "语音服务未配置"}), 503

    data = request.json or {}
    text = _sanitize_spoken_text(data.get('text', ''))
    if not text:
        return jsonify({"status": "error", "message": "文本不能为空"}), 400

    per = data.get('per', 4115)
    spd = data.get('spd', 5)

    try:
        # aue=6 返回 wav（带 header，浏览器可直接播放）
        audio_bytes = _speech_client.text_to_speech(
            text, per=per, spd=spd, vol=8, aue=6
        )
        from flask import Response
        return Response(audio_bytes, mimetype='audio/wav')
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


if __name__ == '__main__':
    server_host = _get_server_host()
    server_port = _get_server_port()
    debug_enabled = _should_enable_debug_mode()
    _assert_server_bindable(server_host, server_port)
    print(f"ProView API 服务监听地址: {_build_server_url(server_host, server_port)}")
    print(f"ProView API 调试模式: {'ON' if debug_enabled else 'OFF'}")
    # 禁用 reloader 避免与 Playwright 冲突
    app.run(debug=debug_enabled, host=server_host, port=server_port, use_reloader=False, threaded=True)
